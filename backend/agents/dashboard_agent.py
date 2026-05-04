"""
Dashboard Agent - Agent #4
Generates QA Activity Dashboard from TFS queries and Excel reports.
Supports two modes:
  - static : tables + charts data only (no LLM call)
  - ai     : static data + LLM strategic analysis narrative
"""

import base64
import io
import logging
import os
from collections import Counter
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

import requests

from ..tfs_tool import (
    sanitize_params,
    _get_auth_and_headers
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# TFS helpers
# ---------------------------------------------------------------------------

def _tfs_get(url: str, auth: Any, headers: dict, timeout: int = 30):
    """Standardized GET for Dashboard Agent with centralized auth support."""
    resp = requests.get(url, auth=auth, headers=headers, timeout=timeout, verify=False)
    resp.raise_for_status()
    return resp.json()


def fetch_tfs_saved_queries(project_url: str, pat: str = None, username: str = None, password: str = None, api_version: str = "5.0") -> List[Dict]:
    """
    Return a flat list of saved queries sorted by lastModifiedDate desc.
    Uses parallel sub-folder fetching so the full tree loads in 1-2 round trips.
    """
    import concurrent.futures

    auth, headers = _get_auth_and_headers(username=username, password=password, pat=pat)
    results: List[Dict] = []
    pending_folders: List[str] = []    # folder paths that need a separate fetch

    def _collect(node: dict):
        if not node:
            return
        if node.get("isFolder"):
            children = node.get("children")
            if children is not None:
                for child in children:
                    _collect(child)
            elif node.get("hasChildren") and node.get("path"):
                # Mark for parallel fetch — don't recurse synchronously
                pending_folders.append(node["path"].lstrip("/"))
        else:
            q_id = node.get("id", "")
            if q_id:
                results.append({
                    "id": q_id,
                    "name": node.get("name", "Unnamed"),
                    "url": node.get("url", ""),
                    "path": node.get("path", ""),
                    "lastModifiedDate": node.get("lastModifiedDate", ""),
                    "createdDate": node.get("createdDate", ""),
                })

    def _fetch_folder(folder_path: str = "") -> dict:
        path_seg = f"/{folder_path}" if folder_path else ""
        url = f"{project_url}/_apis/wit/queries{path_seg}?$depth=2&$expand=all&api-version={api_version}"
        return _tfs_get(url, auth, headers, timeout=12)

    # --- Round 1: fetch root (gets top-level folders + their depth-2 children) ---
    try:
        root_data = _fetch_folder()
        top_nodes = root_data.get("value", []) if "value" in root_data else [root_data]
        for node in top_nodes:
            _collect(node)
    except Exception as exc:
        logger.warning(f"Dashboard: root query fetch failed: {exc}")
        return []

    # --- Round 2: fetch any sub-folders not inlined, all in parallel ---
    if pending_folders:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(8, len(pending_folders))) as pool:
            futures = {pool.submit(_fetch_folder, fp): fp for fp in pending_folders}
            for fut in concurrent.futures.as_completed(futures, timeout=15):
                fp = futures[fut]
                try:
                    data = fut.result()
                    sub_nodes = data.get("value", []) if "value" in data else [data]
                    for node in sub_nodes:
                        _collect(node)
                except Exception as exc:
                    logger.warning(f"Dashboard: sub-folder '{fp}' fetch failed: {exc}")

    results.sort(key=lambda q: q.get("lastModifiedDate") or q.get("createdDate") or "", reverse=True)
    return results


def fetch_work_items_for_query(project_url: str, pat: str = None, username: str = None, password: str = None, query_id: str = "", api_version: str = "5.0") -> List[Dict]:
    """Run a saved WIQL query and return work-item detail rows."""
    auth, headers = _get_auth_and_headers(username=username, password=password, pat=pat)
    wiql_url = f"{project_url}/_apis/wit/wiql/{query_id}?api-version={api_version}"
    try:
        data = _tfs_get(wiql_url, auth, headers)
    except Exception as exc:
        logger.warning(f"Dashboard: WIQL query {query_id} failed: {exc}")
        return []

    ids = [str(item["id"]) for item in data.get("workItems", []) if "id" in item]
    if not ids:
        return []

    all_items: List[Dict] = []
    batch = 150
    fields = (
        "System.Id,System.Title,System.WorkItemType,System.State,"
        "System.AssignedTo,System.CreatedDate,System.ChangedDate,"
        "Microsoft.VSTS.Common.Priority"
    )
    for i in range(0, len(ids), batch):
        chunk = ids[i : i + batch]
        url = (
            f"{project_url}/_apis/wit/workitems"
            f"?ids={','.join(chunk)}&fields={fields}&api-version={api_version}"
        )
        try:
            resp = _tfs_get(url, auth, headers)
            all_items.extend(resp.get("value", []))
        except Exception as exc:
            logger.warning(f"Dashboard: batch fetch failed: {exc}")
    return all_items


# ---------------------------------------------------------------------------
# Excel report parsers  (matching weekly-status automation zip exactly)
# ---------------------------------------------------------------------------

def _normalize(v: Any) -> str:
    return " ".join(str(v or "").strip().lower().replace("\n", " ").split())


def _to_num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_header_row(ws, required: List[str]):
    for row_idx in range(1, min(ws.max_row, 15) + 1):
        values = [_normalize(cell.value) for cell in ws[row_idx]]
        if all(any(req in cell for cell in values) for req in required):
            return row_idx, values
    return None, []


def _parse_vertical_report(file_bytes: bytes) -> Optional[Dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header_row, headers = _find_header_row(ws, ["vertical", "question", "similarity"])
        if not header_row:
            return None
        hm = {h: i for i, h in enumerate(headers)}
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            v = row[hm.get("vertical", -1)] if "vertical" in hm else None
            if not v:
                continue
            chat_name_key = next((k for k in hm if "chatname" in k or "chat name" in k), None)
            asst_key = next((k for k in hm if "assistantname" in k or "assistant name" in k), None)
            q_key = next((k for k in hm if "no of question" in k or "question" in k), None)
            sim_key = next((k for k in hm if "similarity" in k), None)
            rows.append({
                "chat_name": str(row[hm[chat_name_key]] or "") if chat_name_key else "",
                "assistant_name": str(row[hm[asst_key]] or "") if asst_key else "",
                "vertical": str(v),
                "question_count": int(_to_num(row[hm[q_key]] if q_key else 0)),
                "similarity": _to_num(row[hm[sim_key]] if sim_key else 0),
            })
        if not rows:
            return None
        avg_sim = round(sum(r["similarity"] for r in rows) / len(rows), 1)
        top = max(rows, key=lambda x: x["similarity"])
        return {
            "rows": rows,
            "summary": {
                "total_verticals": len(rows),
                "total_questions": sum(r["question_count"] for r in rows),
                "avg_similarity": avg_sim,
                "top_vertical": top["vertical"],
                "top_similarity": top["similarity"],
            },
            "chart": {
                "labels": [r["vertical"] for r in rows],
                "similarity": [r["similarity"] for r in rows],
                "questions": [r["question_count"] for r in rows],
            },
        }
    except Exception as exc:
        logger.warning(f"Dashboard: vertical report parse failed: {exc}")
        return None


def _parse_automation_report(file_bytes: bytes) -> Optional[Dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header_row, headers = _find_header_row(ws, ["module name", "automated test cases", "coverage"])
        if not header_row:
            return None
        hm = {h: i for i, h in enumerate(headers)}
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            mod = row[hm.get("module name", -1)] if "module name" in hm else None
            if not mod:
                continue
            cov_key = next((k for k in hm if "coverage" in k), None)
            auto_key = next((k for k in hm if "automated test cases" in k), None)
            tfs_key = next((k for k in hm if "total test cases" in k and "tfs" in k), None)
            status_key = next((k for k in hm if k == "status"), None)
            rows.append({
                "module_name": str(mod),
                "total_tfs": int(_to_num(row[hm[tfs_key]] if tfs_key else 0)),
                "automated": int(_to_num(row[hm[auto_key]] if auto_key else 0)),
                "coverage": round(_to_num(row[hm[cov_key]] if cov_key else 0), 1),
                "status": str(row[hm[status_key]] or "") if status_key else "",
            })
        if not rows:
            return None
        non_total = [r for r in rows if r["module_name"].strip().lower() != "total"] or rows
        overall_row = next((r for r in rows if r["module_name"].strip().lower() == "total"), None)
        automated_total = overall_row["automated"] if overall_row else sum(r["automated"] for r in non_total)
        tfs_total = overall_row["total_tfs"] if overall_row else sum(r["total_tfs"] for r in non_total)
        overall_coverage = round((automated_total / tfs_total) * 100, 1) if tfs_total else 0.0
        top_modules = sorted(non_total, key=lambda x: x["coverage"], reverse=True)[:8]
        return {
            "rows": rows,
            "summary": {
                "module_count": len(non_total),
                "tfs_total": tfs_total,
                "automated_total": automated_total,
                "overall_coverage": overall_coverage,
            },
            "chart": {
                "labels": [r["module_name"] for r in top_modules],
                "coverage": [r["coverage"] for r in top_modules],
            },
        }
    except Exception as exc:
        logger.warning(f"Dashboard: automation report parse failed: {exc}")
        return None


def _parse_performance_report(file_bytes: bytes) -> Optional[Dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        header_row, headers = _find_header_row(ws, ["test scenario", "users/load", "avg. response time"])
        if not header_row:
            return None
        hm = {h: i for i, h in enumerate(headers)}
        rows = []
        current_scenario = ""
        current_priority = ""
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            sc_key = next((k for k in hm if "test scenario" in k), None)
            scenario = row[hm[sc_key]] if sc_key else None
            if scenario:
                current_scenario = str(scenario)
            if not current_scenario:
                continue
            pri_key = next((k for k in hm if k == "priority"), None)
            priority = row[hm[pri_key]] if pri_key else None
            if priority:
                current_priority = str(priority)
            ul_key = next((k for k in hm if "users/load" in k), None)
            avg_key = next((k for k in hm if "avg. response time" in k), None)
            peak_key = next((k for k in hm if "peak response time" in k), None)
            err_key = next((k for k in hm if "error rate" in k), None)
            res_key = next((k for k in hm if "test result" in k or "pass/fail" in k), None)
            users_load = _to_num(row[hm[ul_key]] if ul_key else 0)
            avg_response = _to_num(row[hm[avg_key]] if avg_key else 0)
            peak_response = _to_num(row[hm[peak_key]] if peak_key else 0)
            error_rate = _to_num(row[hm[err_key]] if err_key else 0)
            result_val = str(row[hm[res_key]] or "") if res_key else ""
            if users_load == 0 and avg_response == 0 and peak_response == 0 and not result_val:
                continue
            rows.append({
                "scenario": current_scenario,
                "priority": current_priority,
                "users_load": users_load,
                "avg_response": avg_response,
                "peak_response": peak_response,
                "error_rate": error_rate,
                "result": result_val,
            })
        if not rows:
            return None
        scenario_max: Dict[str, float] = {}
        for row in rows:
            scenario_max[row["scenario"]] = max(scenario_max.get(row["scenario"], 0), row["peak_response"])
        failure_count = sum(1 for r in rows if r["result"].strip().lower() == "fail")
        worst = max(rows, key=lambda x: x["peak_response"])
        return {
            "rows": rows,
            "summary": {
                "scenario_count": len({r["scenario"] for r in rows}),
                "run_count": len(rows),
                "failure_count": failure_count,
                "worst_peak_ms": worst["peak_response"],
                "worst_scenario": worst["scenario"],
            },
            "chart": {
                "labels": list(scenario_max.keys()),
                "peak_response": [scenario_max[k] for k in scenario_max],
            },
        }
    except Exception as exc:
        logger.warning(f"Dashboard: performance report parse failed: {exc}")
        return None


# ---------------------------------------------------------------------------
# Chart / table builders (mirrors WSA data_processor.py)
# ---------------------------------------------------------------------------

def _classify(state: str) -> str:
    s = (state or "").lower()
    if s in {"closed", "resolved", "done", "completed", "fixed", "verified"}:
        return "closed"
    if s in {"active", "in progress", "in-progress", "new", "open", "approved", "committed"}:
        return "active"
    return "other"


def _state_summary(label: str, items: List[Dict], query_total: Optional[int] = None) -> Dict:
    c = Counter(_classify(it.get("fields", {}).get("System.State", "")) for it in items)
    return {
        "category": label,
        "total": int(query_total if query_total is not None else len(items)),
        "closed": int(c["closed"]),
        "active": int(c["active"]),
        "other": int(c["other"]),
    }


def _assignee_chart(items: List[Dict]) -> Dict:
    counts: Counter = Counter()
    for it in items:
        a = it.get("fields", {}).get("System.AssignedTo", {})
        name = a.get("displayName", "Unassigned") if isinstance(a, dict) else (str(a) if a else "Unassigned")
        counts[name] += 1
    srt = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    return {"labels": [n for n, _ in srt], "values": [int(c) for _, c in srt]}


def _priority_chart(items: List[Dict]) -> Dict:
    _map = {"1": "1 - Critical", "2": "2 - High", "3": "3 - Medium", "4": "4 - Low"}
    counts: Counter = Counter()
    for it in items:
        p = str(it.get("fields", {}).get("Microsoft.VSTS.Common.Priority", "") or "")
        counts[_map.get(p, p or "Unknown")] += 1
    srt = sorted(counts.items())
    return {"labels": [n for n, _ in srt], "values": [int(c) for _, c in srt]}


def _state_chart(items: List[Dict]) -> Dict:
    counts: Counter = Counter()
    for it in items:
        counts[str(it.get("fields", {}).get("System.State", "Unknown"))] += 1
    srt = sorted(counts.items(), key=lambda x: x[1], reverse=True)
    return {"labels": [n for n, _ in srt], "values": [int(c) for _, c in srt]}


def _detailed_table(items: List[Dict]) -> List[Dict]:
    rows = []
    for it in items:
        f = it.get("fields", {})
        a = f.get("System.AssignedTo", {})
        assignee = a.get("displayName", "Unassigned") if isinstance(a, dict) else (str(a) if a else "Unassigned")
        rows.append({
            "id": str(f.get("System.Id", "")),
            "title": str(f.get("System.Title", "")),
            "type": str(f.get("System.WorkItemType", "")),
            "state": str(f.get("System.State", "")),
            "assignee": str(assignee),
            "priority": str(f.get("Microsoft.VSTS.Common.Priority", "")),
            "created": str(f.get("System.CreatedDate", ""))[:10] if f.get("System.CreatedDate") else "",
            "changed": str(f.get("System.ChangedDate", ""))[:10] if f.get("System.ChangedDate") else "",
        })
    return rows


def _category_charts(items: List[Dict]) -> Dict:
    return {
        "assignee": _assignee_chart(items),
        "priority": _priority_chart(items),
        "state": _state_chart(items),
        "table": _detailed_table(items),
    }


def _management_cards(summary: Dict, vertical_report, automation_report, performance_report) -> List[Dict]:
    cards: List[Dict] = [
        {"label": "Total QA Items",      "value": summary["total"],     "tone": "primary"},
        {"label": "Bugs Created By QA",  "value": summary["bugs"],      "tone": "danger"},
        {"label": "Bugs Retested",       "value": summary["retesting"], "tone": "accent"},
        {"label": "User Stories In QA",  "value": summary["stories"],   "tone": "success"},
    ]
    if summary["other"]:
        cards.append({"label": "Other Items", "value": summary["other"], "tone": "primary"})
    if vertical_report:
        cards.append({"label": "Avg Vertical Similarity",
                      "value": f'{vertical_report["summary"]["avg_similarity"]}%', "tone": "success"})
    if automation_report:
        cards.append({"label": "Automation Coverage",
                      "value": f'{automation_report["summary"]["overall_coverage"]}%', "tone": "accent"})
    if performance_report:
        fc = performance_report["summary"]["failure_count"]
        cards.append({"label": "Performance Failures", "value": fc,
                      "tone": "danger" if fc else "success"})
    return cards


# ---------------------------------------------------------------------------
# Default LLM prompt
# ---------------------------------------------------------------------------

DEFAULT_DASHBOARD_PROMPT = """### ROLE:
Act as a Senior Strategic QA Director (20+ years experience). Provide a high-impact, executive-level strategic analysis of the project's quality health.

### DATASETS FOR ANALYSIS:
- TFS ACTIVITY SUMMARY: {tfs_summary}
- VERTICAL VALIDATION REPORT: {vertical_report}
- AUTOMATION COVERAGE REPORT: {automation_report}
- PERFORMANCE REPORT: {performance_report}

### OUTPUT STRUCTURE & STYLE:
Your response MUST be professional and highly readable.
- USE **BOLD CAPITALIZED HEADINGS** for main sections.
- USE *Italicized Title Case* for sub-headings.
- USE standard bullets (• or -) or numbers (1, 2, 3) for points.
- DO NOT use "## **" or "### **" combinations.

1. **EXECUTIVE SUMMARY**
   Provide a 2-sentence high-level summary. Start with a clear "Quality Status" (e.g., EXCELLENT, STABLE, AT RISK, CRITICAL).

2. **KEY QUALITY INDICATORS (KQIs)**
   *Correlation Analysis*
   Briefly interpret the correlation between TFS activity, vertical validation, and automation.

3. **TOP STRATEGIC RISKS**
   *Risk 1: [Title]*
   Root Cause analysis and Potential Impact based on the data.
   *Risk 2: [Title]*
   Root Cause analysis and Potential Impact.

4. **ACTIONABLE ROADMAP**
   *Prioritized Actions*
   Provide 3 high-priority recommendations with clear owners.

5. **CONFIDENCE SCORE**
   Rate from 0 to 100. Provide a one-sentence justification.

Keep the tone professional, authoritative, and focused on delivery excellence."""


# ---------------------------------------------------------------------------
# Main execute function
# ---------------------------------------------------------------------------

def execute_dashboard_agent(
    project_url: str,
    tfs_config: Dict,
    bug_query_id: str = "",
    retest_query_id: str = "",
    story_query_id: str = "",
    other_query_id: str = "",
    vertical_excel_bytes: Optional[bytes] = None,
    automation_excel_bytes: Optional[bytes] = None,
    performance_excel_bytes: Optional[bytes] = None,
    mode: str = "static",
    llm_prompt: str = "",
    llm_config: Optional[Dict] = None,
    api_version: str = "5.0",
) -> Dict:
    """
    Main entry point for Dashboard Agent.
    """
    logger.info(f"🚀 Dashboard Agent execution started (mode={mode})")
    logger.info(f"📊 Inputs: Queries(Bug={bug_query_id}, Retest={retest_query_id}, Story={story_query_id}, Other={other_query_id})")
    logger.info(f"📂 Excel sizes: Vertical={len(vertical_excel_bytes) if vertical_excel_bytes else 0} bytes, "
                f"Automation={len(automation_excel_bytes) if automation_excel_bytes else 0} bytes, "
                f"Performance={len(performance_excel_bytes) if performance_excel_bytes else 0} bytes")

    pat = tfs_config.get("pat_token")
    user = tfs_config.get("username")
    pw = tfs_config.get("password")

    # ---- 1. Fetch work items (Parallelized for massive speedup) ----
    import concurrent.futures
    
    query_map = {
        "Bugs": bug_query_id,
        "Retesting": retest_query_id,
        "Stories": story_query_id,
        "Other": other_query_id
    }
    
    fetch_results = {}
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        future_to_label = {
            executor.submit(fetch_work_items_for_query, project_url, pat, user, pw, qid, api_version): label 
            for label, qid in query_map.items() if qid
        }
        
        for future in concurrent.futures.as_completed(future_to_label):
            label = future_to_label[future]
            try:
                items = future.result()
                fetch_results[label] = items
                logger.info(f"✅ Fetched {len(items)} items for {label}")
            except Exception as exc:
                logger.error(f"❌ Dashboard: parallel fetch for {label} failed: {exc}")
                fetch_results[label] = []

    bug_items    = fetch_results.get("Bugs", [])
    retest_items = fetch_results.get("Retesting", [])
    story_items  = fetch_results.get("Stories", [])
    other_items  = fetch_results.get("Other", [])

    # ---- 2. Summary ------------------------------------------------
    summary = {
        "bugs":      len(bug_items),
        "retesting": len(retest_items),
        "stories":   len(story_items),
        "other":     len(other_items),
        "total":     len(bug_items) + len(retest_items) + len(story_items) + len(other_items),
    }
    logger.info(f"📈 TFS Summary: {summary}")

    # ---- 3. State table --------------------------------------------
    state_rows = [
        _state_summary("Bugs Created By QA",  bug_items,    len(bug_items)),
        _state_summary("Bugs Retested",        retest_items, len(retest_items)),
        _state_summary("User Stories In QA",   story_items,  len(story_items)),
        _state_summary("Other",                other_items,  len(other_items)),
    ]

    # ---- 4. Distribution chart (doughnut) --------------------------
    trend = {
        "labels": ["Bugs Created By QA", "Bugs Retested", "User Stories In QA", "Other"],
        "values": [len(bug_items), len(retest_items), len(story_items), len(other_items)],
    }

    # ---- 5. Per-category charts + tables ---------------------------
    bug_charts    = _category_charts(bug_items)
    retest_charts = _category_charts(retest_items)
    story_charts  = _category_charts(story_items)
    other_charts  = _category_charts(other_items)

    # ---- 6. Excel reports ------------------------------------------
    logger.info("📄 Parsing Excel reports...")
    vertical_report    = None
    if vertical_excel_bytes:
        logger.info("Parsing Vertical report...")
        vertical_report = _parse_vertical_report(vertical_excel_bytes)
        logger.info(f"Vertical report result: {'✅ Success' if vertical_report else '❌ Failed/Empty'}")

    automation_report  = None
    if automation_excel_bytes:
        logger.info("Parsing Automation report...")
        automation_report = _parse_automation_report(automation_excel_bytes)
        logger.info(f"Automation report result: {'✅ Success' if automation_report else '❌ Failed/Empty'}")

    performance_report = None
    if performance_excel_bytes:
        logger.info("Parsing Performance report...")
        performance_report = _parse_performance_report(performance_excel_bytes)
        logger.info(f"Performance report result: {'✅ Success' if performance_report else '❌ Failed/Empty'}")

    # ---- 7. Management cards ---------------------------------------
    mgmt_cards = _management_cards(summary, vertical_report, automation_report, performance_report)

    # ---- 8. Source links (TFS query browser URLs) ------------------
    def _query_url(qid: str) -> Optional[str]:
        if not qid or not project_url:
            return None
        return f"{project_url}/_queries/query/{qid}"

    sources = {
        "bugs":      {"source_url": _query_url(bug_query_id)}    if bug_query_id    else None,
        "retesting": {"source_url": _query_url(retest_query_id)} if retest_query_id else None,
        "stories":   {"source_url": _query_url(story_query_id)}  if story_query_id  else None,
        "other":     {"source_url": _query_url(other_query_id)}  if other_query_id  else None,
    }

    result: Dict = {
        "status":             "success",
        "mode":               mode,
        "summary":            summary,
        "state_rows":         state_rows,
        "trend":              trend,
        "bug_charts":         bug_charts,
        "retest_charts":      retest_charts,
        "story_charts":       story_charts,
        "other_charts":       other_charts,
        "vertical_report":    vertical_report,
        "automation_report":  automation_report,
        "performance_report": performance_report,
        "management_cards":   mgmt_cards,
        "sources":            sources,
        "has_tfs_data":       bool(bug_items or retest_items or story_items or other_items),
        "ai_analysis":        None,
    }

    # ---- 8. AI analysis (only for "ai" mode) -----------------------
    if mode == "ai" and llm_config:
        try:
            from ..llm_config import get_configured_llm
            llm = get_configured_llm(llm_config)
            prompt_template = llm_prompt.strip() or DEFAULT_DASHBOARD_PROMPT
            tfs_txt = (f"Bugs:{summary['bugs']} Retesting:{summary['retesting']} "
                       f"Stories:{summary['stories']} Other:{summary['other']} Total:{summary['total']}")
            prompt = (prompt_template
                      .replace("{tfs_summary}", tfs_txt)
                      .replace("{vertical_report}", str(vertical_report["summary"]) if vertical_report else "No data")
                      .replace("{automation_report}", str(automation_report["summary"]) if automation_report else "No data")
                      .replace("{performance_report}", str(performance_report["summary"]) if performance_report else "No data"))
            result["ai_analysis"] = llm.call([{"role": "user", "content": prompt}])
        except Exception as exc:
            logger.error(f"Dashboard AI analysis failed: {exc}", exc_info=True)
            result["ai_analysis"] = f"AI analysis unavailable: {exc}"

    return result
