# Bulk Task Processing Module
import io
import openpyxl
import requests as req
from typing import Optional, Dict
from datetime import datetime

async def process_bulk_tasks(exec_id: str, file_url: str, sheet_name: str, iteration_path: str, tfs_config: Dict, active_executions: Dict, execution_history: Dict):
    """Process multiple tasks from Excel file"""
    try:
        from .agents.tfs_task_agent import execute_task_creation
        
        active_executions[exec_id]["progress"] = 10
        
        # Download and parse file
        file_data = None
        
        if "drive.google.com" in file_url:
            file_id = file_url.split("/d/")[1].split("/")[0]
            download_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            response = req.get(download_url, timeout=30)
            file_data = io.BytesIO(response.content)
        elif "onedrive.live.com" in file_url or "sharepoint.com" in file_url:
            download_url = file_url.replace("?web=1", "").replace("redir?", "download?") + "&download=1"
            response = req.get(download_url, timeout=30)
            file_data = io.BytesIO(response.content)
        else:
            with open(file_url, "rb") as f:
                file_data = io.BytesIO(f.read())
        
        active_executions[exec_id]["progress"] = 20
        
        # Parse Excel
        workbook = openpyxl.load_workbook(file_data, data_only=True)
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        
        # Extract tasks from Excel
        tasks = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            # Assuming format: [Task Name, Description, ...]
            task_name = row[0]
            task_desc = row[1] if len(row) > 1 else ""
            
            if task_name:
                tasks.append({
                    "name": str(task_name),
                    "description": str(task_desc) if task_desc else "",
                    "row": row_idx
                })
        
        active_executions[exec_id]["total_tasks"] = len(tasks)
        active_executions[exec_id]["progress"] = 30
        
        # Create tasks
        created = []
        failed = []
        
        for idx, task in enumerate(tasks):
            try:
                # Call task creation for each task
                result = execute_task_creation(
                    work_item_id=None,
                    task_description=task["description"],
                    llm_config=None,
                    tfs_config=tfs_config,
                    batch_mode=True,
                    iteration_path=iteration_path,
                    task_title=task["name"]
                )
                
                if result.get("status") == "success":
                    created.append({"task": task["name"], "result": result})
                else:
                    failed.append({"task": task["name"], "error": result.get("message")})
            except Exception as task_error:
                failed.append({"task": task["name"], "error": str(task_error)})
            
            # Update progress
            progress = 30 + int((idx + 1) / len(tasks) * 70)
            active_executions[exec_id]["progress"] = progress
            active_executions[exec_id]["completed"] = len(created)
            active_executions[exec_id]["failed"] = len(failed)
        
        active_executions[exec_id]["status"] = "completed"
        active_executions[exec_id]["progress"] = 100
        active_executions[exec_id]["result"] = {
            "status": "success" if len(failed) == 0 else "partial",
            "created": len(created),
            "failed": len(failed),
            "tasks_created": created,
            "tasks_failed": failed,
            "message": f"Created {len(created)} tasks, {len(failed)} failed"
        }
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        
        # Add to history
        if "default" not in execution_history:
            execution_history["default"] = []
        execution_history["default"].append({
            "execution_id": exec_id,
            "agent": "TFS Bulk Task Creator",
            "status": "completed",
            "timestamp": datetime.now().isoformat(),
            "tasks_created": len(created),
            "tasks_failed": len(failed)
        })
    
    except Exception as e:
        active_executions[exec_id]["status"] = "error"
        active_executions[exec_id]["error"] = str(e)
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        print(f"❌ Error in bulk task creation: {str(e)}")
        import traceback
        traceback.print_exc()
