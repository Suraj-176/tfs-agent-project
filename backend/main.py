from fastapi import FastAPI, HTTPException, BackgroundTasks, Request, UploadFile, File, Form, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel
from typing import Optional, Dict, List
import asyncio
import json
import threading
import logging
from pathlib import Path
from datetime import datetime, timedelta
import io
import re
import base64
import os
import uuid
import time
import tempfile
import math
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import requests
import openpyxl
import csv

from .llm_config import LLMConfigManager, get_configured_llm

# Configure logging to file and console
import sys
from pathlib import Path
from logging.handlers import TimedRotatingFileHandler

log_dir = Path(__file__).parent.parent / "logs"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / "backend.log"

_log_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
_file_handler = TimedRotatingFileHandler(
    log_file, when="midnight", interval=1, backupCount=1, encoding="utf-8"
)
_file_handler.setFormatter(_log_formatter)
_console_handler = logging.StreamHandler(sys.stderr)
_console_handler.setFormatter(_log_formatter)

logging.basicConfig(
    level=logging.INFO,
    handlers=[_file_handler, _console_handler]
)

logger = logging.getLogger(__name__)

# ==================== Cleanup Utilities ====================

def cleanup_old_files(retention_hours: int = 24):
    """Delete files older than retention_hours from logs directory and its subdirectories"""
    try:
        logs_dir = Path(__file__).parent.parent / "logs"
        if not logs_dir.exists():
            return
        
        now = time.time()
        retention_seconds = retention_hours * 3600
        deleted_count = 0
        
        # 1. Clean up files directly in the logs directory (including iis_stdout, bug_process, etc.)
        for file_path in logs_dir.iterdir():
            if file_path.is_file():
                # Skip the active log file
                if file_path.name == "backend.log":
                    continue
                
                file_age = now - file_path.stat().st_mtime
                if file_age > retention_seconds:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        logger.debug(f"🗑️ Deleted old log file: {file_path.name}")
                    except Exception as e:
                        logger.warning(f"Could not delete {file_path}: {e}")

        # 2. Clean up screenshots sub-folder
        screenshots_dir = logs_dir / "screenshots"
        if screenshots_dir.exists():
            for file_path in screenshots_dir.iterdir():
                if file_path.is_file():
                    file_age = now - file_path.stat().st_mtime
                    if file_age > retention_seconds:
                        try:
                            file_path.unlink()
                            deleted_count += 1
                            logger.debug(f"🗑️ Deleted old screenshot: {file_path.name}")
                        except Exception as e:
                            logger.warning(f"Could not delete {file_path}: {e}")
        
        if deleted_count > 0:
            logger.info(f"🧹 Cleanup: Deleted {deleted_count} old files (retention: {retention_hours}h)")
    except Exception as e:
        logger.error(f"Error during cleanup_old_files: {e}")

def schedule_cleanup():
    """Schedule periodic cleanup every hour"""
    def cleanup_loop():
        while True:
            try:
                time.sleep(3600)  # Run every hour
                cleanup_old_files(retention_hours=24)
            except Exception as e:
                logger.error(f"Error in cleanup loop: {e}")
    
    cleanup_thread = threading.Thread(target=cleanup_loop, daemon=True)
    cleanup_thread.start()


# ==================== Configuration Validation ====================

def validate_tfs_config(base_url: str = "", username: str = "", password: str = "", pat: str = "") -> tuple[bool, str]:
    """
    Validate TFS configuration, return human-readable warnings
    Returns: (is_valid, error_message_or_empty_string)
    Note: Returns True even if some fields are missing - they might be provided by env vars
    """
    base_url = (base_url or "").strip()
    username = (username or "").strip()
    password = (password or "").strip()
    pat = (pat or "").strip()
    
    # Just warn if auth is missing, don't block
    if not pat and not (username and password):
        return True, "⚠️ No auth provided - will try environment variables"  # Still valid, just warning
    
    # Log successful validation
    auth_method = "PAT" if pat else "NTLM"
    logger.debug(f"✅ TFS config provided: (Auth: {auth_method})")
    return True, ""


def validate_llm_config(llm_config: dict = None) -> tuple[bool, str]:
    """
    Validate LLM configuration, provide helpful error messages
    Returns: (is_valid, error_message_or_empty_string)
    """
    if not llm_config:
        logger.debug("ℹ️ No LLM config provided, will use default or skip LLM features")
        return True, ""  # Not required
    
    provider = (llm_config.get("provider") or "").lower().strip()
    
    if not provider:
        return False, "❌ LLM provider not specified. Choose: azure, openai, claude, or gemini"
    
    if provider == "azure":
        required_fields = ["deployment_name", "api_version", "endpoint", "api_key"]
        missing = [f for f in required_fields if not (llm_config.get(f) or "").strip()]
        if missing:
            return False, f"❌ Azure LLM missing fields: {', '.join(missing)}"
    elif provider in ["openai", "claude", "gemini"]:
        if not (llm_config.get("api_key") or "").strip():
            return False, f"❌ {provider.upper()} API key not provided"
    else:
        return False, f"❌ Unknown LLM provider: {provider}. Choose: azure, openai, claude, or gemini"
    
    logger.info(f"✅ LLM config validated: provider={provider}")
    return True, ""


def validate_request_config(tfs_config: dict = None, llm_config: dict = None) -> dict:
    """
    Validate configs and log warnings but DON'T BLOCK execution
    Agents should still run even if some config is missing (they use env vars as fallback)
    Returns: dict with validation_passed (always True) and messages (warnings only)
    """
    messages = []
    
    if tfs_config:
        is_valid, msg = validate_tfs_config(
            base_url=tfs_config.get("base_url", ""),
            username=tfs_config.get("username", ""),
            password=tfs_config.get("password", ""),
            pat=tfs_config.get("pat_token", "")
        )
        if msg:  # Just log warnings, don't use is_valid to block
            logger.debug(f"TFS Config: {msg}")
    
    if llm_config:
        is_valid, msg = validate_llm_config(llm_config)
        if msg:
            logger.debug(f"LLM Config: {msg}")
    
    # Always return validation_passed=True to not block execution
    # Agents have fallback logic for missing configs
    return {
        "validation_passed": True,
        "messages": messages,
        "has_warnings": len(messages) > 0
    }


def _safe_json_content(value):
    """Best-effort conversion to JSON-safe content to avoid serialization 500s."""
    def _clean_non_finite(obj):
        if isinstance(obj, float):
            if math.isnan(obj) or math.isinf(obj):
                return None
            return obj
        if isinstance(obj, dict):
            return {k: _clean_non_finite(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_clean_non_finite(v) for v in obj]
        if isinstance(obj, tuple):
            return [_clean_non_finite(v) for v in obj]
        return obj

    try:
        return jsonable_encoder(_clean_non_finite(value))
    except Exception:
        try:
            return jsonable_encoder(_clean_non_finite({"status": "error", "error": str(value)}))
        except Exception:
            return {"status": "error", "error": "Response serialization failed"}

# Initialize FastAPI app
app = FastAPI(title="TFS Agent Hub", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
# ==================== Static Files & Frontend ==================== 

frontend_path = Path(__file__).parent.parent / "frontend"
css_path = frontend_path / "css"
js_path = frontend_path / "js"

# ==================== Models ====================

class LLMConfigRequest(BaseModel):
    provider: str = "azure"
    deployment_name: Optional[str] = None
    api_version: Optional[str] = None
    endpoint: Optional[str] = None
    api_key: Optional[str] = None
    model: Optional[str] = None

class TFSConfigRequest(BaseModel):
    base_url: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    pat_token: Optional[str] = None
    task_url: Optional[str] = None
    test_plan_url: Optional[str] = None

class TaskCreationRequest(BaseModel):
    work_item_id: Optional[int] = None
    task_description: Optional[str] = ""
    tfs_config: Optional[TFSConfigRequest] = None
    batch_mode: Optional[bool] = False
    iteration_path: Optional[str] = None
    excel_file: Optional[str] = None
    sheet_name: Optional[str] = None
    llm_config: Optional[LLMConfigRequest] = None

class TestCaseGenerationRequest(BaseModel):
    work_item_id: Optional[int] = None
    story_details: Optional[str] = None
    sop_text: Optional[str] = ""
    test_mode: Optional[str] = "functional"
    functional_prompt: Optional[str] = ""
    ui_prompt: Optional[str] = ""
    ui_screenshot_name: Optional[str] = ""
    ui_screenshot_data: Optional[str] = ""
    ui_screenshot_names: Optional[List[str]] = None
    ui_screenshot_data_list: Optional[List[str]] = None
    tfs_config: Optional[TFSConfigRequest] = None
    llm_config: Optional[LLMConfigRequest] = None
    coverage_analysis: Optional[bool] = False

class BugCreationRequest(BaseModel):
    """Request for creating bugs or features in TFS"""
    work_item_id: Optional[int] = None
    work_item_type: str = "Bug" # Bug or Feature
    bug_title: str
    bug_description: Optional[str] = ""
    reproduction_steps: Optional[str] = ""
    expected_behavior: Optional[str] = ""
    actual_behavior: Optional[str] = ""
    severity: Optional[str] = "2 - High"
    priority: Optional[str] = "1"
    assigned_to: Optional[str] = None
    found_in_version: Optional[str] = ""
    is_update: bool = False
    tfs_config: Optional[TFSConfigRequest] = None
    llm_config: Optional[LLMConfigRequest] = None

class UserStoryFetchRequest(BaseModel):
    work_item_id: int
    tfs_config: Optional[TFSConfigRequest] = None

class CodeReviewRequest(BaseModel):
    content: str
    review_type: Optional[str] = "general"

class DriveLinkValidationRequest(BaseModel):
    provider: str
    file_url: str
    access_token: Optional[str] = None


class DriveBulkTaskRequest(BaseModel):
    provider: str
    file_url: str
    access_token: Optional[str] = None
    iteration_path: str
    sheet_name: Optional[str] = None
    tfs_config: Optional[TFSConfigRequest] = None
    mode: Optional[str] = "create"  # Added mode parameter

class ExcelDownloadRequest(BaseModel):
    report_rows: List[Dict]
    filename: Optional[str] = "tfs_tasks_report.xlsx"


class OAuthPollRequest(BaseModel):
    session_id: str

class StoryChatRequest(BaseModel):
    story_text: str
    question: str
    chat_history: Optional[List[Dict]] = None
    llm_config: Optional[LLMConfigRequest] = None
    tfs_config: Optional[TFSConfigRequest] = None

class TestCaseAnalysisRequest(BaseModel):
    test_cases: str
    story_details: Optional[str] = ""
    question: str
    chat_history: Optional[List[Dict]] = None
    llm_config: Optional[LLMConfigRequest] = None

class TestCaseReviewRequest(BaseModel):
    test_cases: str
    story_details: Optional[str] = ""
    llm_config: Optional[LLMConfigRequest] = None

class StoryAnalysisRequest(BaseModel):
    story_text: str
    tfs_config: Optional[TFSConfigRequest] = None
    llm_config: Optional[LLMConfigRequest] = None

# ==================== Dashboard Agent Models ====================

class DashboardQueriesRequest(BaseModel):
    tfs_config: TFSConfigRequest

class DashboardGenerateRequest(BaseModel):
    tfs_config: TFSConfigRequest
    llm_config: Optional[LLMConfigRequest] = None
    bug_query_id: Optional[str] = ""
    retest_query_id: Optional[str] = ""
    story_query_id: Optional[str] = ""
    other_query_id: Optional[str] = ""
    # Base64-encoded Excel file bytes
    vertical_excel_b64: Optional[str] = None
    automation_excel_b64: Optional[str] = None
    performance_excel_b64: Optional[str] = None
    mode: Optional[str] = "static"      # "static" or "ai"
    llm_prompt: Optional[str] = ""

# ==================== Routes & Initialization ====================

@app.get("/", include_in_schema=False)
async def serve_index():
    """Serve the main frontend page at the root URL."""
    index_file = frontend_path / "index.html"
    if index_file.exists():
        return FileResponse(index_file)
    return JSONResponse(status_code=404, content={"message": "Frontend index.html not found"})

# Mount static files
if css_path.exists():
    app.mount("/css", StaticFiles(directory=str(css_path)), name="css")
if js_path.exists():
    app.mount("/js", StaticFiles(directory=str(js_path)), name="js")

# Mount templates for direct download
templates_dir = Path(__file__).parent.parent / "templates"
if templates_dir.exists():
    app.mount("/templates", StaticFiles(directory=str(templates_dir)), name="templates")
# We don't mount the root as static to avoid shadowing the API routes

@app.post("/api/agent/tfs-task/download-excel")
async def download_task_result_excel(request: ExcelDownloadRequest):
    """Generate and return an Excel file from task report rows."""
    try:
        from .agents.tfs_task_agent import generate_task_excel_report
        
        excel_bytes = generate_task_excel_report(request.report_rows)
        
        filename = request.filename or "tfs_tasks_report.xlsx"
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
            
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error generating Excel: {e}")
        return JSONResponse(status_code=500, content={"message": f"Error generating Excel: {str(e)}"})

# ==================== Startup & Shutdown ====================

@app.on_event("startup")
async def startup_event():
    """Run cleanup and schedule periodic cleanup on app startup"""
    logger.info("🚀 Application starting...")
    # Clean up old files immediately on startup
    cleanup_old_files(retention_hours=24)
    # Schedule periodic cleanup
    schedule_cleanup()
    logger.info("✅ Cleanup scheduler initialized")

# Store for execution tracking (session-based, in-memory)
execution_history: Dict[str, List[Dict]] = {}
active_executions: Dict[str, Dict] = {}
oauth_device_sessions: Dict[str, Dict] = {}
chat_states: Dict[str, Dict] = {}


def _is_probable_csv(file_bytes: bytes) -> bool:
    try:
        sample = file_bytes[:4096].decode("utf-8", errors="ignore")
        lines = [line for line in sample.splitlines() if line.strip()]
        if not lines:
            return False
        return ("," in lines[0] or "\t" in lines[0]) and len(lines) >= 1
    except Exception:
        return False


def _extract_google_file_id(url: str) -> Optional[str]:
    match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)

    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    ids = query.get("id")
    if ids:
        return ids[0]

    return None


def _build_download_candidates(provider: str, file_url: str) -> List[str]:
    provider = (provider or "").strip().lower()
    url = (file_url or "").strip()
    candidates: List[str] = []

    if provider == "gdrive":
        file_id = _extract_google_file_id(url)

        # Google Sheets URL -> export directly as xlsx
        if "docs.google.com/spreadsheets/" in url and file_id:
            candidates.append(f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx")

        if file_id:
            candidates.append(f"https://drive.google.com/uc?export=download&id={file_id}")

        candidates.append(url)

    elif provider == "onedrive":
        # OneDrive/SharePoint direct download attempts
        if "download=1" in url:
            candidates.append(url)
        else:
            joiner = "&" if "?" in url else "?"
            candidates.append(f"{url}{joiner}download=1")

        # SharePoint Doc.aspx style link -> download.aspx variants
        if "/_layouts/15/Doc.aspx" in url:
            parsed = urlparse(url)
            q = parse_qs(parsed.query)

            # Variant 1: replace Doc.aspx with download.aspx, preserve existing query.
            dl_path = parsed.path.replace("/_layouts/15/Doc.aspx", "/_layouts/15/download.aspx")
            candidates.append(urlunparse((parsed.scheme, parsed.netloc, dl_path, "", parsed.query, "")))

            # Variant 2: explicit action=download for some tenants.
            q2 = {k: v[:] for k, v in q.items()}
            q2["action"] = ["download"]
            candidates.append(urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode(q2, doseq=True), "")))

            # Variant 3: download.aspx + action=download.
            candidates.append(urlunparse((parsed.scheme, parsed.netloc, dl_path, "", urlencode(q2, doseq=True), "")))

            # Variant 4: minimal query for download.aspx (drop mobile/default page flags).
            q3 = {}
            if "sourcedoc" in q:
                q3["sourcedoc"] = q["sourcedoc"]
            if "file" in q:
                q3["file"] = q["file"]
            q3["download"] = ["1"]
            candidates.append(urlunparse((parsed.scheme, parsed.netloc, dl_path, "", urlencode(q3, doseq=True), "")))

        # OneDrive public sharing API form (often avoids 403 on raw share links)
        try:
            encoded = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")
            share_token = f"u!{encoded}"
            candidates.append(f"https://api.onedrive.com/v1.0/shares/{share_token}/root/content")
        except Exception:
            pass

        candidates.append(url)
    else:
        candidates.append(url)

    # Remove duplicates while preserving order
    seen = set()
    deduped = []
    for candidate in candidates:
        if candidate not in seen:
            deduped.append(candidate)
            seen.add(candidate)
    return deduped


def _download_excel_bytes(
    candidates: List[str],
    provider: str = "",
    access_token: str = "",
    source_url: str = "",
) -> tuple[bytes, str]:
    last_error = "Unknown error"

    # Authenticated provider API fallbacks for private links.
    auth_candidates: List[tuple[str, dict]] = []
    if access_token:
        token_headers = {
            "Authorization": f"Bearer {access_token}",
            "User-Agent": "Mozilla/5.0",
            "Accept": "*/*",
        }

        if provider == "onedrive" and source_url:
            try:
                encoded = base64.urlsafe_b64encode(source_url.encode("utf-8")).decode("utf-8").rstrip("=")
                share_token = f"u!{encoded}"
                auth_candidates.append((f"https://graph.microsoft.com/v1.0/shares/{share_token}/driveItem/content", token_headers))
                auth_candidates.append((f"https://api.onedrive.com/v1.0/shares/{share_token}/root/content", token_headers))
            except Exception:
                pass

        if provider == "gdrive" and source_url:
            fid = _extract_google_file_id(source_url)
            if fid:
                if "docs.google.com/spreadsheets/" in source_url:
                    auth_candidates.append((
                        f"https://www.googleapis.com/drive/v3/files/{fid}/export?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        token_headers
                    ))
                auth_candidates.append((f"https://www.googleapis.com/drive/v3/files/{fid}?alt=media", token_headers))

    for auth_url, auth_headers in auth_candidates:
        try:
            auth_res = requests.get(auth_url, timeout=25, allow_redirects=True, headers=auth_headers)
            if auth_res.status_code == 200 and auth_res.content:
                ctype = (auth_res.headers.get("Content-Type") or "").lower()
                if "text/html" not in ctype:
                    return auth_res.content, auth_url
            else:
                last_error = f"HTTP {auth_res.status_code}"
        except Exception as ex:
            last_error = str(ex)

    for candidate in candidates:
        try:
            response = requests.get(
                candidate,
                timeout=25,
                allow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0", "Accept": "*/*"},
            )
            if response.status_code != 200:
                if response.status_code == 403:
                    last_error = "HTTP 403 (link permissions blocked direct download)"
                else:
                    last_error = f"HTTP {response.status_code}"
                continue

            content = response.content or b""
            if not content:
                last_error = "Empty file content"
                continue

            # Some SharePoint links return an HTML landing page first.
            ctype = (response.headers.get("Content-Type") or "").lower()
            if "text/html" in ctype:
                html = response.text or ""
                html_links = _extract_links_from_html(html, candidate)
                if html_links:
                    for html_link in html_links:
                        try:
                            nested_headers = {"User-Agent": "Mozilla/5.0", "Accept": "*/*"}
                            if access_token:
                                nested_headers["Authorization"] = f"Bearer {access_token}"
                            nested = requests.get(
                                html_link,
                                timeout=25,
                                allow_redirects=True,
                                headers=nested_headers,
                            )
                            if nested.status_code == 200 and nested.content:
                                nested_type = (nested.headers.get("Content-Type") or "").lower()
                                if "text/html" not in nested_type:
                                    return nested.content, html_link
                        except Exception:
                            continue

            if len(content) > 25 * 1024 * 1024:
                last_error = "File is too large (max 25 MB for validation)"
                continue

            return content, candidate
        except Exception as ex:
            last_error = str(ex)

    raise ValueError(last_error)


def _extract_links_from_html(html_text: str, base_url: str = "") -> List[str]:
    links: List[str] = []
    if not html_text:
        return links

    # Absolute links
    for m in re.finditer(r"https?://[^\"'\\s<>]+", html_text, flags=re.IGNORECASE):
        url = m.group(0)
        if "download.aspx" in url.lower() or "onedrive.com" in url.lower() or "sharepoint.com" in url.lower():
            links.append(url)

    # Relative download.aspx links
    for m in re.finditer(r"(/[^\"'\\s<>]*download\\.aspx[^\"'\\s<>]*)", html_text, flags=re.IGNORECASE):
        rel = m.group(1)
        if base_url:
            parsed = urlparse(base_url)
            links.append(f"{parsed.scheme}://{parsed.netloc}{rel}")

    # De-dup preserve order
    seen = set()
    out = []
    for link in links:
        if link not in seen:
            seen.add(link)
            out.append(link)
    return out


def _extract_sheet_names(file_bytes: bytes) -> List[str]:
    try:
        sample = file_bytes[:512].decode("utf-8", errors="ignore").strip().lower()
        if sample.startswith("<!doctype html") or sample.startswith("<html"):
            raise ValueError("Received an HTML page instead of a file. The share link is not a direct downloadable file link.")

        # XLSX files are ZIP containers; fast reject HTML/other payloads.
        if file_bytes.startswith(b"PK"):
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            return workbook.sheetnames or []

        if _is_probable_csv(file_bytes):
            return ["Sheet1"]

        # Last attempt for workbook parsers.
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        return workbook.sheetnames or []
    except Exception as ex:
        raise ValueError(f"File is accessible but not a readable Excel workbook: {str(ex)}")


def _oauth_device_config(provider: str) -> dict:
    p = (provider or "").strip().lower()
    if p == "onedrive":
        client_id = os.getenv("ONEDRIVE_OAUTH_CLIENT_ID", "").strip()
        tenant = os.getenv("ONEDRIVE_OAUTH_TENANT_ID", "common").strip() or "common"
        scope = os.getenv("ONEDRIVE_OAUTH_SCOPE", "offline_access Files.Read User.Read").strip()
        return {
            "provider": "onedrive",
            "client_id": client_id,
            "device_url": f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/devicecode",
            "token_url": f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token",
            "scope": scope,
        }

    if p == "gdrive":
        client_id = os.getenv("GOOGLE_OAUTH_CLIENT_ID", "").strip()
        scope = os.getenv(
            "GOOGLE_OAUTH_SCOPE",
            "https://www.googleapis.com/auth/drive.readonly https://www.googleapis.com/auth/spreadsheets.readonly",
        ).strip()
        return {
            "provider": "gdrive",
            "client_id": client_id,
            "device_url": "https://oauth2.googleapis.com/device/code",
            "token_url": "https://oauth2.googleapis.com/token",
            "scope": scope,
        }

    raise ValueError("Unsupported provider. Use 'onedrive' or 'gdrive'.")


def _oauth_session_is_expired(session: dict) -> bool:
    created = session.get("created_at", 0)
    expires_in = session.get("expires_in", 0)
    return (time.time() - created) > max(0, int(expires_in))

# ==================== API Endpoints ====================

@app.get("/")
async def root():
    """Serve the frontend index.html"""
    frontend_file = Path(__file__).parent.parent / "frontend" / "index.html"
    if frontend_file.exists():
        return FileResponse(str(frontend_file))
    return {"message": "Frontend not found"}

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

# ==================== Prompts Endpoints ====================

@app.get("/api/prompts/all")
async def get_all_prompts():
    """Get all available test case generation prompts"""
    from .prompts_manager import PromptsManager
    
    prompts = PromptsManager.load_prompts()
    return {
        "status": "success",
        "prompts": prompts,
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/prompts/functional")
async def get_functional_prompt():
    """Get functional test case prompt"""
    from .prompts_manager import PromptsManager
    
    prompt = PromptsManager.get_functional_prompt()
    return {
        "status": "success",
        "prompt": prompt,
        "type": "functional",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/prompts/ui")
async def get_ui_prompt():
    """Get UI test case prompt"""
    from .prompts_manager import PromptsManager
    
    prompt = PromptsManager.get_ui_prompt()
    return {
        "status": "success",
        "prompt": prompt,
        "type": "ui",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/prompts/combined")
async def get_combined_prompt():
    """Get combined functional + UI test case prompt"""
    from .prompts_manager import PromptsManager
    
    prompt = PromptsManager.get_combined_prompt()
    return {
        "status": "success",
        "prompt": prompt,
        "type": "combined",
        "timestamp": datetime.now().isoformat()
    }

@app.post("/api/prompts/reload")
async def reload_prompts():
    """Reload prompts from file (cache refresh)"""
    from .prompts_manager import PromptsManager
    
    PromptsManager.reload_prompts()
    return {
        "status": "success",
        "message": "Prompts reloaded from file",
        "timestamp": datetime.now().isoformat()
    }

# ==================== LLM Configuration Endpoints ====================

@app.get("/api/llm/config/default")
async def get_default_llm_config():
    """Get default LLM config from .env"""
    config = LLMConfigManager.load_from_env()
    
    # Convert snake_case to camelCase for frontend
    normalized = {
        "provider": config.get("provider"),
        "deploymentName": config.get("deployment_name"),
        "apiVersion": config.get("api_version"),
        "endpoint": config.get("endpoint"),
        "apiKey": config.get("api_key"),
        "model": config.get("model")
    }
    
    return {
        "config": {k: v for k, v in normalized.items() if v},
        "mask_api_key": True
    }

@app.post("/api/llm/config/validate")
async def validate_llm_config_endpoint(config: LLMConfigRequest):
    """Validate LLM configuration"""
    is_valid, message = LLMConfigManager.validate_config(config.dict())
    return {
        "valid": is_valid,
        "message": message,
        "timestamp": datetime.now().isoformat()
    }

# ==================== OAuth Device Flow Endpoints ====================

@app.post("/api/oauth/{provider}/device/start")
async def oauth_device_start(provider: str):
    """Start OAuth device-code flow for OneDrive or Google Drive."""
    try:
        cfg = _oauth_device_config(provider)
        if not cfg.get("client_id"):
            return {
                "success": False,
                "message": f"Missing OAuth client id for {provider}. Configure environment first.",
                "timestamp": datetime.now().isoformat(),
            }

        response = requests.post(
            cfg["device_url"],
            data={
                "client_id": cfg["client_id"],
                "scope": cfg["scope"],
            },
            timeout=25,
        )
        payload = response.json() if response.content else {}
        if response.status_code >= 400:
            return {
                "success": False,
                "message": payload.get("error_description") or payload.get("error") or f"HTTP {response.status_code}",
                "timestamp": datetime.now().isoformat(),
            }

        session_id = str(uuid.uuid4())
        oauth_device_sessions[session_id] = {
            "provider": cfg["provider"],
            "client_id": cfg["client_id"],
            "token_url": cfg["token_url"],
            "device_code": payload.get("device_code"),
            "interval": int(payload.get("interval", 5)),
            "expires_in": int(payload.get("expires_in", 900)),
            "created_at": time.time(),
        }

        return {
            "success": True,
            "session_id": session_id,
            "provider": cfg["provider"],
            "verification_uri": payload.get("verification_uri") or payload.get("verification_url"),
            "verification_uri_complete": payload.get("verification_uri_complete"),
            "user_code": payload.get("user_code"),
            "interval": int(payload.get("interval", 5)),
            "expires_in": int(payload.get("expires_in", 900)),
            "message": payload.get("message") or "Open verification URL and enter user code.",
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "message": f"Failed to start OAuth flow: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/oauth/{provider}/device/poll")
async def oauth_device_poll(provider: str, request: OAuthPollRequest):
    """Poll OAuth device-code flow status."""
    try:
        session = oauth_device_sessions.get(request.session_id)
        if not session:
            return {"success": False, "status": "error", "message": "OAuth session not found"}

        if session.get("provider") != provider:
            return {"success": False, "status": "error", "message": "Provider/session mismatch"}

        if _oauth_session_is_expired(session):
            oauth_device_sessions.pop(request.session_id, None)
            return {"success": False, "status": "expired", "message": "OAuth session expired"}

        token_data = {
            "client_id": session["client_id"],
            "device_code": session["device_code"],
            "grant_type": "urn:ietf:params:oauth:grant-type:device_code",
        }

        response = requests.post(session["token_url"], data=token_data, timeout=25)
        payload = response.json() if response.content else {}

        if response.status_code == 200 and payload.get("access_token"):
            oauth_device_sessions.pop(request.session_id, None)
            return {
                "success": True,
                "status": "approved",
                "access_token": payload.get("access_token"),
                "refresh_token": payload.get("refresh_token"),
                "expires_in": payload.get("expires_in"),
                "scope": payload.get("scope"),
                "token_type": payload.get("token_type", "Bearer"),
            }

        err = (payload.get("error") or "").lower()
        if err in {"authorization_pending", "slow_down"}:
            return {"success": True, "status": "pending", "error": err}
        if err == "expired_token":
            oauth_device_sessions.pop(request.session_id, None)
            return {"success": False, "status": "expired", "message": "Device code expired"}

        return {
            "success": False,
            "status": "error",
            "message": payload.get("error_description") or payload.get("error") or f"HTTP {response.status_code}",
        }
    except Exception as e:
        return {"success": False, "status": "error", "message": f"OAuth poll failed: {str(e)}"}

# ==================== TFS Integration Endpoints ====================

@app.post("/api/tfs/authenticate")
async def authenticate_tfs(request: TFSConfigRequest):
    """Authenticate with TFS server - only test connection"""
    try:
        import requests
        
        base_url = request.base_url
        username = request.username
        password = request.password
        pat = request.pat_token
        
        if "/tfs/" not in base_url:
            return {
                "success": False,
                "message": "Invalid TFS URL format. Expected: http://server:8080/tfs/Collection/Project",
                "timestamp": datetime.now().isoformat()
            }
        
        headers = {}
        auth_obj = None
        if pat:
            import base64
            headers = {"Authorization": f"Basic {base64.b64encode(f':{pat}'.encode()).decode()}"}
        elif username and password:
            # Prefer NTLM for Windows auth, keep Basic as fallback for compatibility.
            try:
                from requests_ntlm import HttpNtlmAuth
                auth_obj = HttpNtlmAuth(username, password)
            except Exception:
                import base64
                headers = {"Authorization": f"Basic {base64.b64encode(f'{username}:{password}'.encode()).decode()}"}

        # Smart collection URL derivation: we need /tfs/CollectionName/_apis/projects
        # If user provides http://server:8080/tfs/DefaultCollection/ProjectName, 
        # we extract http://server:8080/tfs/DefaultCollection
        parts = base_url.rstrip('/').split('/')
        try:
            tfs_idx = next(i for i, p in enumerate(parts) if p.lower() == "tfs")
            if len(parts) >= tfs_idx + 2:
                collection_url = "/".join(parts[:tfs_idx+2])
                test_url = collection_url + "/_apis/projects"
            else:
                test_url = base_url.rstrip('/') + "/_apis/projects"
        except (StopIteration, Exception):
            test_url = base_url.rstrip('/') + "/_apis/projects"

        response = requests.get(test_url, auth=auth_obj, headers=headers, timeout=10)
        # Fallback path: some servers reject NTLM handshake from this endpoint but accept Basic.
        if response.status_code == 401 and auth_obj is not None and username and password:
            import base64
            basic_headers = {"Authorization": f"Basic {base64.b64encode(f'{username}:{password}'.encode()).decode()}"}
            response = requests.get(test_url, headers=basic_headers, timeout=10)
        if response.status_code in [200, 203]:
            return {
                "success": True,
                "message": "Successfully connected to TFS server",
                "authenticated": True,
                "timestamp": datetime.now().isoformat()
            }
        elif response.status_code == 401:
            return {
                "success": False,
                "message": "Authentication failed (401). Please verify username/password or PAT token.",
                "authenticated": False,
                "timestamp": datetime.now().isoformat()
            }
        else:
            return {
                "success": False,
                "message": f"Failed to connect to TFS: HTTP {response.status_code}",
                "timestamp": datetime.now().isoformat()
            }
    except Exception as e:
        print(f"Error authenticating TFS: {str(e)}")
        return {
            "success": False,
            "message": f"Failed to authenticate with TFS: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }


@app.post("/api/tfs/test-create-permission")
async def test_tfs_create_permission(request: TFSConfigRequest):
    """Validate whether provided credentials can access Task create endpoint."""
    try:
        from .tfs_tool import create_task

        title = f"Permission Validation {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        response = create_task(
            title=title,
            base_url=request.base_url,
            pat=request.pat_token,
            username=request.username,
            password=request.password,
            validate_only=True,
        )

        if response.status_code in [200, 201]:
            return {
                "success": True,
                "can_create": True,
                "message": "Create permission check passed.",
                "timestamp": datetime.now().isoformat(),
            }

        if response.status_code == 400:
            # validateOnly can still return 400 for field/rule issues, which means auth reached the endpoint.
            return {
                "success": True,
                "can_create": True,
                "message": "Authenticated to create endpoint (validation returned 400 due field/rule constraints).",
                "timestamp": datetime.now().isoformat(),
            }

        if response.status_code == 401:
            return {
                "success": False,
                "can_create": False,
                "message": "Authentication failed (401).",
                "timestamp": datetime.now().isoformat(),
            }

        if response.status_code == 403:
            return {
                "success": False,
                "can_create": False,
                "message": "Authenticated but no permission to create tasks (403).",
                "timestamp": datetime.now().isoformat(),
            }

        return {
            "success": False,
            "can_create": False,
            "message": f"Create permission check failed: HTTP {response.status_code}",
            "raw": response.text[:300],
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "can_create": False,
            "message": f"Create permission test failed: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }

@app.post("/api/tfs/fetch-iteration")
async def fetch_iteration_on_demand(request: TFSConfigRequest):
    """Fetch current iteration path from TFS"""
    try:
        from .tfs_tool import fetch_current_iteration
        
        iteration_path = fetch_current_iteration(
            base_url=request.base_url,
            username=request.username,
            password=request.password,
            pat=request.pat_token
        )
        
        if iteration_path:
            return {
                "success": True,
                "iteration_path": iteration_path,
                "message": "Current iteration loaded successfully",
                "timestamp": datetime.now().isoformat()
            }
        else:
            return {
                "success": False,
                "iteration_path": None,
                "message": "Could not fetch iteration. Please enter manually.",
                "timestamp": datetime.now().isoformat()
            }
    except Exception as e:
        print(f"Error fetching iteration: {str(e)}")
        return {
            "success": False,
            "message": f"Failed to fetch iteration: {str(e)}",
            "iteration_path": None,
            "timestamp": datetime.now().isoformat()
        }


@app.post("/api/tfs/iterations")
async def fetch_iteration_list_on_demand(request: TFSConfigRequest):
    """Fetch all available iteration paths from TFS."""
    try:
        from .tfs_tool import fetch_iteration_options

        options = fetch_iteration_options(
            base_url=request.base_url,
            username=request.username,
            password=request.password,
            pat=request.pat_token
        )

        current = next((row.get("path") for row in options if row.get("time_frame") == "current"), None)
        return {
            "success": len(options) > 0,
            "iterations": options,
            "current_iteration": current,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "iterations": [],
            "message": f"Failed to fetch iteration list: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/tfs/areas")
async def fetch_area_list_on_demand(request: TFSConfigRequest):
    """Fetch all available area paths from TFS."""
    try:
        from .tfs_tool import fetch_area_options

        areas = fetch_area_options(
            base_url=request.base_url,
            username=request.username,
            password=request.password,
            pat=request.pat_token
        )

        return {
            "success": len(areas) > 0,
            "areas": areas,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "areas": [],
            "message": f"Failed to fetch area list: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/tfs/tags")
async def fetch_tag_list_on_demand(request: TFSConfigRequest):
    """Fetch all available tags from the TFS project."""
    try:
        from .tfs_tool import _get_auth_and_headers, _normalize_tfs_url_for_api
        import requests
        
        base_url = (request.base_url or "").strip()
        if not base_url:
            return {"success": False, "tags": [], "message": "TFS base URL missing"}
            
        url_base = _normalize_tfs_url_for_api(base_url)
        auth, headers = _get_auth_and_headers(
            username=request.username, 
            password=request.password, 
            pat=request.pat_token
        )
        
        # TFS API for tags: {collection}/{project}/_apis/wit/tags
        tags_url = f"{url_base}/_apis/wit/tags?api-version=6.0"
        response = requests.get(tags_url, auth=auth, headers=headers, timeout=10)
        
        tags = []
        if response.status_code == 200:
            data = response.json()
            # Extract tag names from the response value list
            tags = [tag.get("name") for tag in data.get("value", []) if tag.get("name")]
            
        return {
            "success": True,
            "tags": sorted(tags),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "success": False, 
            "tags": [], 
            "message": f"Failed to fetch tags: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }


@app.post("/api/tfs/work-items")
async def fetch_work_items_for_linking(request: TFSConfigRequest):
    """Fetch recent work items (stories, tasks) for linking to bugs."""
    try:
        from .tfs_tool import _get_auth_and_headers, _normalize_tfs_url_for_api
        import requests
        
        base_url = (request.base_url or "").strip()
        if not base_url:
            return {
                "success": False,
                "work_items": [],
                "message": "TFS base URL not configured",
            }
        
        url_base = _normalize_tfs_url_for_api(base_url)
        auth, headers = _get_auth_and_headers(
            username=request.username,
            password=request.password,
            pat=request.pat_token
        )
        
        # Query for recent work items (Focus ONLY on User Stories as requested)
        wiql_url = f"{url_base}/_apis/wit/wiql?api-version=6.0"
        wiql_body = {
            "query": "SELECT [System.Id], [System.Title], [System.State] FROM WorkItems WHERE [System.WorkItemType] IN ('User Story', 'Story') AND [System.State] <> 'Closed' ORDER BY [System.ChangedDate] DESC"
        }
        
        response = requests.post(wiql_url, auth=auth, headers=headers, json=wiql_body, timeout=10)
        
        work_items = []
        if response.status_code == 200:
            data = response.json()
            ids = [wi['id'] for wi in data.get("workItems", [])[:200]]  # Increased to 200 items
            
            if ids:
                ids_str = ",".join(map(str, ids))
                bulk_url = f"{url_base}/_apis/wit/workitems?ids={ids_str}&fields=System.Title,System.State&api-version=6.0"
                try:
                    bulk_res = requests.get(bulk_url, auth=auth, headers=headers, timeout=15)
                    if bulk_res.status_code == 200:
                        bulk_data = bulk_res.json()
                        # Map items by ID for easy lookup
                        items_map = {item['id']: item for item in bulk_data.get("value", [])}

                        # Re-sort items based on the original IDs order (which was sorted by ChangedDate DESC)
                        for item_id in ids:
                            if item_id in items_map:
                                item = items_map[item_id]
                                fields = item.get("fields", {})
                                work_items.append({
                                    "id": item['id'],
                                    "title": fields.get("System.Title", f"Work Item {item['id']}"),
                                    "state": fields.get("System.State", "Unknown")
                                })

                except Exception as e:
                    logger.error(f"Error in bulk work item fetch: {str(e)}")
        
        return {
            "success": len(work_items) > 0,
            "work_items": work_items,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "work_items": [],
            "message": f"Failed to fetch work items: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


class SuiteRequest(BaseModel):
    project: str
    plan_id: Optional[str] = None
    tfs_config: Optional[TFSConfigRequest] = None


class CreateSuiteRequest(BaseModel):
    suite_name: str
    project: str
    plan_id: Optional[str] = None
    tfs_config: Optional[TFSConfigRequest] = None


class UploadTestCasesRequest(BaseModel):
    test_cases: List[Dict] = []
    suite_id: Optional[str] = None
    work_item_id: Optional[int] = None
    project: str
    plan_id: Optional[str] = None
    tfs_config: Optional[TFSConfigRequest] = None


@app.post("/api/tfs/plans")
async def fetch_plans(request: SuiteRequest):
    """Fetch available test plans for a given project."""
    logger.info(f"📥 FETCH_PLANS CALLED: project={request.project}")
    logger.info(f"📥 TFS config: base_url={request.tfs_config.base_url if request.tfs_config else 'None'}")
    
    try:
        from .tfs_tool import fetch_test_plans
        import asyncio
        from concurrent.futures import ThreadPoolExecutor
        
        if not request.tfs_config:
            logger.error("❌ TFS configuration not provided")
            return {
                "success": False,
                "plans": [],
                "message": "TFS configuration not provided",
                "timestamp": datetime.now().isoformat(),
            }

        logger.info(f"📤 Calling fetch_test_plans in thread pool with:")
        logger.info(f"  - collection_url: {request.tfs_config.base_url}")
        logger.info(f"  - project: {request.project}")
        
        # Run the blocking function in a thread pool to avoid blocking the event loop
        loop = asyncio.get_event_loop()
        plans = await loop.run_in_executor(
            None,  # Use default ThreadPoolExecutor
            fetch_test_plans,
            request.tfs_config.base_url,
            request.project,
            request.tfs_config.username,
            request.tfs_config.password,
            request.tfs_config.pat_token
        )
        
        logger.info(f"✅ fetch_test_plans returned {len(plans)} plans")
        
        return {
            "success": True,
            "plans": plans,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.error(f"❌ Exception in fetch_plans: {str(e)}", exc_info=True)
        return {
            "success": False,
            "plans": [],
            "message": f"Failed to fetch plans: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/tfs/suites")
async def fetch_suites(request: SuiteRequest):
    """Fetch available test suites for a given project and plan."""
    logger.info(f"📥 FETCH_SUITES CALLED: project={request.project}, plan_id={request.plan_id}")
    logger.info(f"📥 TFS config: base_url={request.tfs_config.base_url if request.tfs_config else 'None'}")
    
    try:
        from .tfs_tool import fetch_test_suites
        import asyncio
        
        if not request.tfs_config:
            logger.error("❌ TFS configuration not provided")
            return {
                "success": False,
                "suites": [],
                "message": "TFS configuration not provided",
                "timestamp": datetime.now().isoformat(),
            }

        logger.info(f"📤 Calling fetch_test_suites in thread pool with:")
        logger.info(f"  - collection_url: {request.tfs_config.base_url}")
        logger.info(f"  - project: {request.project}")
        logger.info(f"  - plan_id: {request.plan_id}")
        
        # Run the blocking function in a thread pool to avoid blocking the event loop
        loop = asyncio.get_event_loop()
        suites = await loop.run_in_executor(
            None,  # Use default ThreadPoolExecutor
            fetch_test_suites,
            request.tfs_config.base_url,
            request.project,
            request.plan_id,
            request.tfs_config.username,
            request.tfs_config.password,
            request.tfs_config.pat_token
        )
        
        logger.info(f"✅ fetch_test_suites returned {len(suites)} suites")
        
        return {
            "success": True,
            "suites": suites,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.error(f"❌ Exception in fetch_suites: {str(e)}", exc_info=True)
        return {
            "success": False,
            "suites": [],
            "message": f"Failed to fetch suites: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/tfs/create-suite")
async def create_test_suite(request: CreateSuiteRequest):
    """Create a new test suite in TFS."""
    try:
        from .tfs_upload import create_static_suite
        
        logger.info(f"📥 CREATE_SUITE CALLED")
        logger.info(f"  - suite_name: {request.suite_name}")
        logger.info(f"  - project: {request.project}")
        logger.info(f"  - plan_id: {request.plan_id}")
        logger.info(f"  - tfs_config: {request.tfs_config is not None}")
        
        if not request.tfs_config:
            return {
                "success": False,
                "suite_id": None,
                "message": "TFS configuration not provided",
                "timestamp": datetime.now().isoformat(),
            }

        auth, headers = _get_auth_from_tfs_config(request.tfs_config)
        
        logger.info(f"  - base_url: {request.tfs_config.base_url}")
        logger.info(f"  - Calling create_static_suite...")
        
        suite_id = create_static_suite(
            collection_url=request.tfs_config.base_url,
            project=request.project,
            plan_id=request.plan_id,
            suite_name=request.suite_name,
            auth=auth,
            extra_headers=headers
        )

        logger.info(f"✅ Suite created: {suite_id}")
        return {
            "success": True,
            "suite_id": suite_id,
            "message": f"Suite '{request.suite_name}' created successfully",
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.error(f"❌ Exception in create_test_suite: {str(e)}", exc_info=True)
        return {
            "success": False,
            "suite_id": None,
            "message": f"Failed to create suite: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/tfs/upload-test-cases")
async def upload_test_cases(request: UploadTestCasesRequest):
    """Upload generated test cases to a TFS suite."""
    try:
        from .tfs_upload import create_test_case, add_test_cases_to_suite
        
        logger.info(f"📥 UPLOAD_TEST_CASES CALLED")
        logger.info(f"  - test_cases count: {len(request.test_cases)}")
        logger.info(f"  - suite_id: {request.suite_id}")
        logger.info(f"  - plan_id: {request.plan_id}")
        
        if not request.tfs_config or not request.suite_id:
            logger.warning(f"❌ Missing TFS config or suite ID")
            return {
                "success": False,
                "uploaded": 0,
                "message": "TFS configuration or suite ID not provided",
                "timestamp": datetime.now().isoformat(),
            }

        auth, headers = _get_auth_from_tfs_config(request.tfs_config)
        created_ids = []

        # Create each test case and link to user story
        for i, test_case in enumerate(request.test_cases):
            try:
                logger.info(f"  Creating test case {i+1}/{len(request.test_cases)}: {test_case.get('title', 'Untitled')}")
                
                tc_id = create_test_case(
                    collection_url=request.tfs_config.base_url,
                    project=request.project,
                    title=test_case.get("title", "Test Case"),
                    step_action="\n".join([f"{s.get('action', '')}" for s in test_case.get("steps", [])]),
                    expected_results="\n".join([f"{s.get('expected', '')}" for s in test_case.get("steps", [])]),
                    story_work_item_id=request.work_item_id or 0,
                    auth=auth,
                    extra_headers=headers
                )
                created_ids.append(tc_id)
                logger.info(f"  ✅ Created test case {tc_id}")
            except Exception as e:
                logger.error(f"  ❌ Error creating test case: {str(e)}")
                print(f"Error creating test case '{test_case.get('title')}': {str(e)}")
                continue

        logger.info(f"  Created {len(created_ids)}/{len(request.test_cases)} test cases")

        # Add created test cases to the suite
        if created_ids and request.suite_id:
            try:
                logger.info(f"  Adding {len(created_ids)} test cases to suite {request.suite_id}")
                
                add_test_cases_to_suite(
                    collection_url=request.tfs_config.base_url,
                    project=request.project,
                    plan_id=request.plan_id,
                    suite_id=request.suite_id,
                    test_case_ids=created_ids,
                    auth=auth,
                    extra_headers=headers
                )
                logger.info(f"  ✅ Added test cases to suite")
            except Exception as e:
                logger.warning(f"  ⚠️ Warning: Failed to add test cases to suite: {str(e)}")
                print(f"Warning: Failed to add test cases to suite: {str(e)}")

        logger.info(f"✅ Upload complete: {len(created_ids)} test cases")
        
        return {
            "success": len(created_ids) > 0,
            "uploaded": len(created_ids),
            "created_ids": created_ids,
            "message": f"Uploaded {len(created_ids)} test case(s) to suite",
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.error(f"❌ Exception in upload_test_cases: {str(e)}", exc_info=True)
        return {
            "success": False,
            "uploaded": 0,
            "message": f"Failed to upload test cases: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


def _get_auth_from_tfs_config(tfs_config: TFSConfigRequest):
    """Helper function to extract auth details from TFS config."""
    from .tfs_upload import _get_auth
    return _get_auth(
        pat=tfs_config.pat_token or "",
        username=tfs_config.username or "",
        password=tfs_config.password or ""
    )


@app.post("/api/agent/create-bug")
async def create_bug_endpoint(request: BugCreationRequest):
    """
    Agent #4: Work Item Creation Endpoint
    Creates bug or feature work items in TFS using the Agent
    """
    wi_type = request.work_item_type or "Bug"
    logger.info(f"📝 {wi_type.upper()} CREATION REQUESTED: title={request.bug_title}")
    
    try:
        from .agents.bug_creation_agent import execute_bug_creation
        
        if not request.tfs_config:
            return {
                "success": False,
                "message": "TFS configuration is required",
                "timestamp": datetime.now().isoformat(),
            }
        
        # Execute creation using the agent
        result = execute_bug_creation(
            work_item_id=request.work_item_id,
            work_item_type=wi_type,
            bug_description=request.bug_description,
            bug_title=request.bug_title,
            reproduction_steps=request.reproduction_steps,
            expected_behavior=request.expected_behavior,
            actual_behavior=request.actual_behavior,
            severity=request.severity,
            priority=request.priority,
            llm_config=request.llm_config.model_dump() if request.llm_config else None,
            tfs_config=request.tfs_config.model_dump() if request.tfs_config else None,
            found_in_version=request.found_in_version,
            assigned_to=request.assigned_to,
        )
        
        logger.info(f"✅ {wi_type} creation completed: {request.bug_title}")
        
        # 🔧 AUTO SELF-HEAL: Automatically review and validate bug details
        try:
            from .agents.code_reviewer_agent import execute_code_review
            
            bug_details_for_review = f"""
Title: {request.bug_title}
Description: {request.bug_description}
Reproduction Steps: {request.reproduction_steps}
Expected Behavior: {request.expected_behavior}
Actual Behavior: {request.actual_behavior}
Severity: {request.severity}
Priority: {request.priority}
"""
            review_result = execute_code_review(bug_details_for_review, request.llm_config.model_dump() if request.llm_config else None)
            logger.info(f"🔧 Auto self-heal review completed: {review_result.get('status', 'completed')}")
        except Exception as e:
            logger.debug(f"ℹ️ Self-heal review skipped (optional): {str(e)}")
        
        return {
            "success": result.get("success", False),
            "bug_title": request.bug_title,
            "status": result.get("status", "Created"),
            "analysis": result.get("analysis", ""),
            "message": result.get("message", "Bug creation analysis completed"),
            "timestamp": datetime.now().isoformat(),
        }
    
    except Exception as e:
        logger.error(f"❌ Error in bug creation: {str(e)}", exc_info=True)
        return {
            "success": False,
            "bug_id": None,
            "error": str(e),
            "message": f"Failed to create bug: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/agent/format-bug-report")
async def format_bug_report(request: dict):
    """
    Format bug or feature description into professional report using AI
    """
    try:
        from .prompts_manager import PromptsManager
        
        bug_description = request.get("bug_description", "").strip()
        wi_type = request.get("work_item_type", "Bug")
        logger.info(f"✨ FORMAT {wi_type.upper()} REPORT REQUESTED: {bug_description[:50]}...")
        
        # Support both single and multiple screenshots
        screenshots = request.get("screenshots", [])
        screenshot_file = request.get("screenshot_file")
        screenshot_filename = request.get("screenshot_filename", "screenshot.png")
        
        if screenshot_file and not screenshots:
            screenshots = [{"filename": screenshot_filename, "data": screenshot_file}]
            
        history = request.get("history", [])
        llm_config = request.get("llm_config")
        
        if not bug_description and not screenshots:
            return {
                "success": False,
                "error": f"{wi_type} description or screenshot is required",
                "timestamp": datetime.now().isoformat(),
            }
        
        # Construct enhanced description with all screenshots mentioned
        current_input = bug_description
        if screenshots:
            screenshot_mentions = "\n\n" + "\n".join([f"[Screenshot attached: {s.get('filename', 'screenshot.png')}]" for s in screenshots])
            current_input += screenshot_mentions
        
        # Use LLM to format the report
        from .llm_config import get_configured_llm
        try:
            llm = get_configured_llm(llm_config) if llm_config else get_configured_llm()
            
            # Select prompt based on work item type
            from .prompts_manager import PromptsManager
            if wi_type.lower() == "feature":
                base_prompt = PromptsManager.get_feature_report_prompt() or "Convert this into a professional Feature work item."
            else:
                base_prompt = PromptsManager.get_bug_report_prompt() or "Convert this into a professional Bug work item."
                
            system_prompt = f"""
            {base_prompt}
            
            Your response must be a JSON object containing:
            - title: A clear, concise {wi_type} title
            - description: A professional summary (Overview for Features)
            - reproduction_steps: (For Bugs) Numbered steps to reproduce; (For Features) Requirements
            - expected_behavior: (For Bugs) What should happen; (For Features) Business Value
            - actual_behavior: (For Bugs) What actually happens; (For Features) Acceptance Criteria
            - severity: (For Bugs) One of "1 - Critical", "2 - High", "3 - Medium", "4 - Low"
            - priority: One of "1", "2", "3"
            
            STRICT RULES:
            1. NEVER use "**" or other placeholder markers as content. 
            2. If a section is provided with "**" by the user, it means "fill this in". 
            3. Use the conversation history to summarize what should go there.
            4. If a section is truly unknown, leave it as an empty string.
            
            When creating the 'formatted_report' string, use these EXACT labels:
            Title: ...
            Description: ...
            Steps to Reproduce: ...
            Actual Result: ...
            Expected Result: ...
            (For Features use Business Value:, Requirements:, Acceptance Criteria:)
            """
            
            # Build messages for LLM
            messages = [{"role": "system", "content": system_prompt}]
            # Add history (limit to last 4 for efficiency)
            for h in history[-4:]:
                messages.append(h)
            # Add current user input
            messages.append({"role": "user", "content": f"New info to incorporate into the {wi_type} report: {current_input}"})
            
            response = llm.call(messages)
            content = str(response)
            logger.info(f"🤖 LLM response received (length: {len(content)})")
            
            import json
            import re
            
            # Clean up content: sometimes LLM adds markdown blocks around JSON
            content_cleaned = content.strip()
            if "```json" in content_cleaned:
                content_cleaned = content_cleaned.split("```json")[1].split("```")[0].strip()
            elif "```" in content_cleaned:
                content_cleaned = content_cleaned.split("```")[1].split("```")[0].strip()
                
            structured_data = {}
            success = False
            
            # Try to parse as direct JSON first
            try:
                # Find JSON object
                json_match = re.search(r'\{.*\}', content_cleaned, re.DOTALL)
                if json_match:
                    structured_data = json.loads(json_match.group(0))
                    logger.info(f"✅ Successfully parsed JSON from LLM")
                    success = True
            except Exception as json_e:
                logger.warning(f"⚠️ Failed to parse JSON from LLM: {str(json_e)}")
            
            # Fallback: Extraction using Regex if JSON parsing failed
            if not success:
                logger.info("🔍 Attempting fallback extraction using regex...")
                
                # Use simpler regex patterns that work reliably
                # For multi-line content, use DOTALL flag and greedy matching until next field
                structured_data = {}
                
                # 1. Robust Title Extraction (handle bolding and colons)
                title_match = re.search(r'"title"\s*:\s*"(.*?)"', content, re.I) or \
                              re.search(r'(?:\*\*|#|)\s*Title\s*(?:\*\*|)\s*[:\-]*\s*(.+?)(?:\n|$)', content, re.I)
                
                title_val = title_match.group(1).strip() if title_match else ""
                title_val = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', title_val).strip()
                
                # Fallback Title
                if not title_val or title_val == "**":
                    lines = [l.strip() for l in content.split('\n') if l.strip()]
                    for line in lines:
                        clean_line = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', line).strip()
                        if clean_line and len(clean_line) > 5 and not any(k in clean_line.lower() for k in ["description", "overview", "steps", "actual", "expected"]):
                            title_val = clean_line[:70]
                            break
                    if not title_val: title_val = f"New {wi_type} from Chat"
                structured_data["title"] = title_val
                
                # 2. Extract Description (Overview)
                desc_match = re.search(r'"description"\s*:\s*"(.*?)"', content, re.I) or \
                             re.search(r'(?:\*\*|#|)\s*(?:Description|Overview)\s*(?:\*\*|)\s*[:\-]*\s*(.*?)(?=(?:\*\*|#|)\s*(?:Steps to Reproduce|Reproduction Steps|Steps|Actual Result|Actual Behavior|Actual|Expected Result|Expected Behavior|Expected|Severity|Priority)|$)', content, re.I | re.DOTALL)
                
                desc_val = desc_match.group(1).strip() if desc_match else ""
                desc_val = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', desc_val).strip()
                structured_data["description"] = desc_val
                
                # 3. Extract Reproduction Steps (Requirements for Features)
                # More flexible: try JSON, then try various header formats, then try just "Steps:" anywhere
                steps_match = re.search(r'"(?:reproduction_steps|steps)"\s*:\s*"([^"]*(?:\n[^"]*)*)"', content, re.I) or \
                              re.search(r'"reproduction_steps"\s*:\s*"(.*?)"', content, re.I) or \
                              re.search(r'(?:\*\*|#|)\s*(?:Steps to Reproduce|Reproduction Steps|Steps|Requirements)\s*(?:\*\*|)\s*[:\-]*\s*(.*?)(?=(?:\*\*|#|)\s*(?:Actual Result|Actual Behavior|Actual|Expected Result|Expected Behavior|Expected|Acceptance Criteria|Severity|Priority)|$)', content, re.I | re.DOTALL) or \
                              re.search(r'(?:^|\n)\s*(?:Steps?|Repro|Reproduce)\s*[:\-]*\s*(.*?)(?=(?:^|\n)\s*(?:Actual|Expected|Severity|Priority|Acceptance)|$)', content, re.I | re.DOTALL)
                
                steps_val = steps_match.group(1).strip() if steps_match else ""
                steps_val = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', steps_val).strip()
                structured_data["reproduction_steps"] = steps_val
                
                # 4. Extract Actual Result (Acceptance Criteria for Features)
                actual_match = re.search(r'"(?:actual_behavior|actual_result)"\s*:\s*"([^"]*(?:\n[^"]*)*)"', content, re.I) or \
                               re.search(r'"actual_behavior"\s*:\s*"(.*?)"', content, re.I) or \
                               re.search(r'(?:\*\*|#|)\s*(?:Actual Result|Actual Behavior|Actual|Acceptance Criteria)\s*(?:\*\*|)\s*[:\-]*\s*(.*?)(?=(?:\*\*|#|)\s*(?:Expected Result|Expected Behavior|Expected|Severity|Priority)|$)', content, re.I | re.DOTALL) or \
                               re.search(r'(?:^|\n)\s*(?:Actual|Current\s+Behavior|What\s+Happens)\s*[:\-]*\s*(.*?)(?=(?:^|\n)\s*(?:Expected|Severity|Priority)|$)', content, re.I | re.DOTALL)
                
                actual_val = actual_match.group(1).strip() if actual_match else ""
                actual_val = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', actual_val).strip()
                structured_data["actual_behavior"] = actual_val
                
                # 5. Extract Expected Result (Business Value for Features)
                expected_match = re.search(r'"(?:expected_behavior|expected_result)"\s*:\s*"([^"]*(?:\n[^"]*)*)"', content, re.I) or \
                                 re.search(r'"expected_behavior"\s*:\s*"(.*?)"', content, re.I) or \
                                 re.search(r'(?:\*\*|#|)\s*(?:Expected Result|Expected Behavior|Expected|Business Value)\s*(?:\*\*|)\s*[:\-]*\s*(.*?)(?=(?:\*\*|#|)\s*(?:Severity|Priority)|$)', content, re.I | re.DOTALL) or \
                                 re.search(r'(?:^|\n)\s*(?:Expected|Should\s+(?:Be|Happen)|Desired\s+Behavior)\s*[:\-]*\s*(.*?)(?=(?:^|\n)\s*(?:Severity|Priority)|$)', content, re.I | re.DOTALL)
                
                expected_val = expected_match.group(1).strip() if expected_match else ""
                expected_val = re.sub(r'^[\*\s#:\-]+|[\*\s#:\-]+$', '', expected_val).strip()
                structured_data["expected_behavior"] = expected_val
                
                # Extract Severity and Priority
                severity_match = re.search(r'"severity"\s*:\s*"(.*?)"', content, re.I) or re.search(r'Severity:\s*(.+?)(?:\n|$)', content, re.I)
                structured_data["severity"] = severity_match.group(1).strip() if severity_match else "2 - High"
                
                priority_match = re.search(r'"priority"\s*:\s*"(.*?)"', content, re.I) or re.search(r'Priority:\s*(.+?)(?:\n|$)', content, re.I)
                structured_data["priority"] = priority_match.group(1).strip() if priority_match else "2"
                
                logger.info(f"✅ Fallback extraction completed")
                logger.info(f"   title: {repr(structured_data.get('title', '')[:50])}")
                logger.info(f"   description: {repr(structured_data.get('description', '')[:100])}")
                logger.info(f"   reproduction_steps: {repr(structured_data.get('reproduction_steps', '')[:100])}")
                
                # Check if we at least got a title or description
                if structured_data["title"] or structured_data["description"]:
                    logger.info("✅ Fallback extraction successful")
                    success = True

            if success:
                # Create a formatted string for the UI to display and for app.js to parse
                if wi_type.lower() == "feature":
                    formatted_report = f"""Title: {structured_data.get('title', 'Feature')}
Description: {structured_data.get('description', '')}
Business Value: {structured_data.get('expected_behavior', '')}
Requirements: {structured_data.get('reproduction_steps', '')}
Acceptance Criteria: {structured_data.get('actual_behavior', '')}
Severity: {structured_data.get('severity', '2 - High')}
Priority: {structured_data.get('priority', '2')}
"""
                else:
                    formatted_report = f"""Title: {structured_data.get('title', 'Bug')}
Description: {structured_data.get('description', '')}
Steps to Reproduce: {structured_data.get('reproduction_steps', '')}
Actual Result: {structured_data.get('actual_behavior', '')}
Expected Result: {structured_data.get('expected_behavior', '')}
Severity: {structured_data.get('severity', '2 - High')}
Priority: {structured_data.get('priority', '2')}
"""
                logger.info(f"=== FORMATTED REPORT FOR UI ===")
                logger.info(f"formatted_report (first 300 chars):\n{formatted_report[:300]}")
                return {
                    "success": True,
                    "data": structured_data,
                    "formatted_report": formatted_report,
                    "severity": structured_data.get('severity', '2 - High'),
                    "priority": structured_data.get('priority', '2'),
                    "timestamp": datetime.now().isoformat()
                }
            else:
                logger.warning(f"⚠️ All parsing attempts failed for LLM response: {content[:300]}...")
                return {
                    "success": False,
                    "error": f"Failed to parse {wi_type} report. Please try again with a clearer description.",
                    "message": f"Failed to format {wi_type} report: Parsing Error. Response: {content[:100]}",
                    "timestamp": datetime.now().isoformat(),
                }
        except Exception as llm_error:
            logger.error(f"❌ LLM error during formatting: {str(llm_error)}", exc_info=True)
            return {
                "success": False,
                "error": str(llm_error),
                "message": f"Failed to format {wi_type} report: LLM Error: {str(llm_error)}",
                "timestamp": datetime.now().isoformat(),
            }
            
        return {
            "success": False,
            "message": f"Failed to format {wi_type} report using AI (Unexpected flow)",
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.error(f"❌ Exception in format_bug_report: {str(e)}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "message": f"Failed to format: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


def ensure_bug_report_format(report_text: str) -> str:
    """Ensure bug report has proper formatting with line breaks between sections"""
    lines = report_text.split('\n')
    formatted_lines = []
    
    section_headers = ['Title:', 'Description:', 'Steps to Reproduce:', 'Actual Result:', 'Expected Result:']
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        
        # Check if this line is a section header
        is_header = any(stripped.startswith(header) for header in section_headers)
        
        if is_header:
            # Add blank line before section if not at start and previous line not blank
            if formatted_lines and formatted_lines[-1].strip() != '':
                formatted_lines.append('')
            
            # Extract header and content if on same line
            for header in section_headers:
                if stripped.startswith(header):
                    content_after_header = stripped[len(header):].strip()
                    formatted_lines.append(header)
                    if content_after_header:
                        formatted_lines.append(content_after_header)
                    break
        elif stripped:  # Non-empty, non-header line
            formatted_lines.append(line)
    
    # Join with newlines and clean up multiple blank lines
    result = '\n'.join(formatted_lines)
    
    # Replace multiple blank lines with single blank line
    while '\n\n\n' in result:
        result = result.replace('\n\n\n', '\n\n')
    
    return result.strip()


def extract_bug_report_sections(formatted_report: str) -> dict:
    """
    Extract individual sections from a formatted bug report
    
    Args:
        formatted_report: Full formatted report with sections
        
    Returns:
        dict with keys: title, description, steps_to_reproduce, actual_result, expected_result
    """
    sections = {
        'title': '',
        'description': '',
        'steps_to_reproduce': '',
        'actual_result': '',
        'expected_result': '',
        'full_report': formatted_report
    }
    
    # Normalize the report format first
    report = ensure_bug_report_format(formatted_report)
    
    # Split into lines
    lines = report.split('\n')
    
    current_section = None
    section_content = []
    
    section_mapping = {
        'Title:': 'title',
        'Description:': 'description',
        'Steps to Reproduce:': 'steps_to_reproduce',
        'Actual Result:': 'actual_result',
        'Expected Result:': 'expected_result'
    }
    
    for line in lines:
        # Check if this is a section header
        is_section_header = False
        for header, key in section_mapping.items():
            if line.strip().startswith(header):
                # Save previous section
                if current_section and section_content:
                    sections[current_section] = '\n'.join(section_content).strip()
                
                # Start new section
                current_section = key
                section_content = []
                
                # Extract content after header if on same line
                content_after_header = line.strip()[len(header):].strip()
                if content_after_header:
                    section_content.append(content_after_header)
                
                is_section_header = True
                break
        
        # If not a header and we have a current section, add to content
        if not is_section_header and current_section and line.strip():
            section_content.append(line)
    
    # Save the last section
    if current_section and section_content:
        sections[current_section] = '\n'.join(section_content).strip()
    
    logger.info(f"📄 Extracted sections from formatted report:")
    logger.info(f"   Title: {sections['title'][:50]}..." if len(sections['title']) > 50 else f"   Title: {sections['title']}")
    logger.info(f"   Description: {sections['description'][:50]}..." if sections['description'] else "   Description: (empty)")
    logger.info(f"   Steps: {sections['steps_to_reproduce'][:50]}..." if sections['steps_to_reproduce'] else "   Steps: (empty)")
    
    return sections


@app.post("/api/tfs/team-members")
async def search_team_members(request: dict):
    """Fetch real TFS team members with integrated search."""
    try:
        from .tfs_tool import search_tfs_identities, BASE_URL, USERNAME, PASSWORD, PAT, _get_auth_and_headers, _split_collection_and_project, get_current_user
        import requests
        
        search_query = request.get("search_query", "").strip()
        tfs_config = request.get("tfs_config")
        
        base_url = (tfs_config or {}).get("base_url", BASE_URL)
        username = (tfs_config or {}).get("username", USERNAME)
        password = (tfs_config or {}).get("password", PASSWORD)
        pat = (tfs_config or {}).get("pat_token", PAT)
        
        if not base_url:
            return {"success": False, "error": "TFS base URL not configured", "members": []}
            
        collection_base, _ = _split_collection_and_project(base_url)
        if not collection_base: collection_base = base_url

        session = requests.Session()
        auth, headers = _get_auth_and_headers(username=username, password=password, pat=pat)
        session.auth = auth
        session.headers.update(headers or {})

        members = []
        seen = set()

        # 0. Add current user as the first option
        try:
            curr = get_current_user(base_url=base_url, username=username, password=password, pat=pat)
            if curr.get("success"):
                user_id = curr.get("id")
                display_name = curr.get("display_name")
                email = curr.get("email")
                if user_id and user_id not in seen:
                    seen.add(user_id)
                    members.append({
                        "id": user_id,
                        "display_name": f"{display_name} (Me)",
                        "email": email or f"{display_name.lower().replace(' ', '.')}@example.com"
                    })
        except Exception as e:
            logger.debug(f"Failed to get current user: {e}")

        # 1. Use identity search if query exists
        if search_query and len(search_query) >= 2:
            identities = search_tfs_identities(name_query=search_query, base_url=base_url, pat=pat, username=username, password=password)
            if identities:
                for id_str in identities:
                    if id_str in seen: continue
                    seen.add(id_str)
                    
                    display_name = id_str
                    email = ""
                    if "<" in id_str and ">" in id_str:
                        parts = id_str.split("<")
                        display_name = parts[0].strip()
                        email = parts[1].replace(">", "").strip()
                    else:
                        display_name = id_str.split("\\")[-1] if "\\" in id_str else id_str
                    
                    members.append({
                        "id": id_str,
                        "display_name": display_name,
                        "email": email or f"{display_name.lower().replace(' ', '.')}@example.com"
                    })
                return {"success": True, "members": members}

        # 2. General fetching fallback (WIQL for recent users)
        try:
            wiql_url = f"{base_url.rstrip('/')}/_apis/wit/wiql?api-version=6.0"
            wiql_body = {"query": "SELECT [System.Id], [System.AssignedTo] FROM WorkItems WHERE [System.AssignedTo] <> '' ORDER BY [System.ChangedDate] DESC"}
            
            response = session.post(wiql_url, json=wiql_body, timeout=10)
            if response.status_code == 200:
                data = response.json()
                for workitem in data.get("workItems", [])[:30]:
                    item_url = f"{collection_base}/_apis/wit/workitems/{workitem['id']}?fields=System.AssignedTo&api-version=6.0"
                    item_res = session.get(item_url, timeout=5)
                    if item_res.status_code == 200:
                        user = item_res.json().get("fields", {}).get("System.AssignedTo")
                        if isinstance(user, dict): user = user.get("displayName")
                        user_str = str(user).strip() if user else ""
                        if user_str and user_str not in seen:
                            seen.add(user_str)
                            display_name = user_str
                            email = ""
                            if "<" in user_str and ">" in user_str:
                                parts = user_str.split("<")
                                display_name = parts[0].strip()
                                email = parts[1].replace(">", "").strip()
                            else:
                                display_name = user_str.split("\\")[-1] if "\\" in user_str else user_str
                            
                            members.append({"id": user_str, "display_name": display_name, "email": email or f"{display_name.lower().replace(' ', '.')}@example.com"})
            return {"success": True, "members": members}
        except Exception as e:
            logger.warning(f"Fallback search failed: {e}")
            return {"success": True, "members": members}
    except Exception as e:
        logger.error(f"Error in search_team_members: {e}")
        return {"success": False, "error": str(e), "members": []}

@app.post("/api/tfs/search-identities")
async def search_tfs_identities_endpoint(request: dict):
    """Search for TFS identities matching a name query."""
    try:
        from .tfs_tool import search_tfs_identities, BASE_URL, USERNAME, PASSWORD, PAT
        search_query = request.get("search_query", "").strip()
        tfs_config = request.get("tfs_config")
        base_url = (tfs_config or {}).get("base_url", BASE_URL)
        username = (tfs_config or {}).get("username", USERNAME)
        password = (tfs_config or {}).get("password", PASSWORD)
        pat = (tfs_config or {}).get("pat_token", PAT)
        
        if not search_query: return {"success": True, "identities": []}
            
        identities = search_tfs_identities(name_query=search_query, base_url=base_url, pat=pat, username=username, password=password)
        return {"success": True, "identities": identities}
    except Exception as e:
        logger.error(f"Error in search-identities: {str(e)}")
        return {"success": False, "error": str(e), "identities": []}


@app.post("/api/agent/create-bug-tfs")
async def create_bug_tfs(request: dict):
    """
    Create or update a bug/feature in TFS with all fields
    
    Request body:
    {
        "bug_title": "Title",
        "description": "Description/Overview",
        "reproduction_steps": "Steps to reproduce",
        "expected_behavior": "Expected result or acceptance criteria",
        "actual_behavior": "Actual result or business value",
        "priority": "1-3",
        "severity": "1-4",
        "tags": "comma-separated",
        "assigned_to": "DOMAIN\\username",
        "related_work_item_id": "123",
        "work_item_id": "12345", (for updates)
        "work_item_type": "Bug" or "Feature",
        "is_update": true/false,
        "screenshots": [{"filename": "file.png", "data": "base64 data"}, ...],
        "tfs_config": {...},
        "llm_config": {...},
        "area_path": "...",
        "iteration_path": "..."
    }
    """
    try:
        from .tfs_tool import create_bug, BASE_URL, USERNAME, PASSWORD, PAT
        from .prompts_manager import PromptsManager
        import logging
        logger = logging.getLogger(__name__)
        
        bug_title = request.get("bug_title", "").strip()
        description = request.get("description", "").strip()
        reproduction_steps = request.get("reproduction_steps", "").strip()
        expected_behavior = request.get("expected_behavior", "").strip()
        actual_behavior = request.get("actual_behavior", "").strip()
        priority = request.get("priority", "2")
        severity = request.get("severity", "2 - High")
        tags = request.get("tags", "").strip()
        assigned_to = request.get("assigned_to")
        related_work_item_id = request.get("related_work_item_id")
        work_item_id = request.get("work_item_id")
        work_item_type = request.get("work_item_type", "Bug")
        is_update = request.get("is_update", False)
        screenshots = request.get("screenshots", [])
        tfs_config = request.get("tfs_config")
        llm_config = request.get("llm_config")
        area_path = request.get("area_path", "").strip()
        iteration_path = request.get("iteration_path", "").strip()
        
        # LOG: What the endpoint received from frontend
        logger.info(f"=== ENDPOINT RECEIVED FROM FRONTEND ===")
        logger.info(f"bug_title: {repr(bug_title[:50] if bug_title else 'EMPTY')}")
        logger.info(f"description: {repr(description[:100] if description else 'EMPTY')}")
        logger.info(f"reproduction_steps: {repr(reproduction_steps[:100] if reproduction_steps else 'EMPTY')}")
        logger.info(f"expected_behavior: {repr(expected_behavior[:100] if expected_behavior else 'EMPTY')}")
        logger.info(f"actual_behavior: {repr(actual_behavior[:100] if actual_behavior else 'EMPTY')}")
        logger.info(f"is_update: {is_update}")
        logger.info(f"work_item_id: {work_item_id}")
        
        # Sanitize placeholders
        for placeholder in ["Search area...", "Search iteration...", "Loading...", "Search..."]:
            if area_path == placeholder: area_path = ""
            if iteration_path == placeholder: iteration_path = ""
        
        # Validation for create
        if not is_update:
            if not bug_title:
                return {"success": False, "error": f"{work_item_type} title is required"}
            if not reproduction_steps and not description:
                return {"success": False, "error": "Description/Steps are required"}
        else:
            if not work_item_id:
                return {"success": False, "error": "Work Item ID is required for update"}
        
        # Get TFS config
        base_url = (tfs_config or {}).get("base_url", BASE_URL)
        username = (tfs_config or {}).get("username", USERNAME)
        password = (tfs_config or {}).get("password", PASSWORD)
        pat = (tfs_config or {}).get("pat_token", PAT)
        
        logger.info(f"📋 Create/Update Work Item called (Type: {work_item_type}, Update: {is_update})")
        logger.info(f"=== CALLING EXECUTE_BUG_CREATION ===")
        logger.info(f"  bug_title: {repr(bug_title[:50] if bug_title else 'EMPTY')}")
        logger.info(f"  bug_description: {repr(description[:100] if description else 'EMPTY')}")
        logger.info(f"  reproduction_steps: {repr(reproduction_steps[:100] if reproduction_steps else 'EMPTY')}")
        logger.info(f"  expected_behavior: {repr(expected_behavior[:100] if expected_behavior else 'EMPTY')}")
        logger.info(f"  actual_behavior: {repr(actual_behavior[:100] if actual_behavior else 'EMPTY')}")
        
        from .agents.bug_creation_agent import execute_bug_creation
        
        result = execute_bug_creation(
            work_item_id=work_item_id,
            related_work_item_id=related_work_item_id,
            work_item_type=work_item_type,
            bug_title=bug_title,
            bug_description=description, 
            reproduction_steps=reproduction_steps,
            expected_behavior=expected_behavior,
            actual_behavior=actual_behavior,
            severity=severity,
            priority=priority,
            tags=tags,
            assigned_to=assigned_to,
            llm_config=llm_config,
            tfs_config=tfs_config or {
                "base_url": base_url,
                "pat_token": pat,
                "username": username,
                "password": password
            },
            area_path=area_path,
            iteration_path=iteration_path,
            is_update=is_update,
            screenshots=screenshots
        )
        
        return result
    
    except Exception as e:
        logger.error(f"❌ Error in create_bug_tfs: {str(e)}", exc_info=True)
        return {"success": False, "error": str(e)}


@app.post("/api/tfs/fetch-bug-details")
async def fetch_bug_details_endpoint(request: dict):
    """
    Fetch existing bug details from TFS by bug ID
    
    Request body:
    {
        "bug_id": 12345,
        "tfs_config": {...}
    }
    """
    try:
        from .tfs_tool import fetch_bug_details
        
        bug_id = request.get("bug_id")
        tfs_config = request.get("tfs_config")
        
        if not bug_id:
            return {
                "success": False,
                "error": "Bug ID is required",
                "timestamp": datetime.now().isoformat()
            }
        
        if not tfs_config:
            return {
                "success": False,
                "error": "TFS configuration is required",
                "timestamp": datetime.now().isoformat()
            }
        
        logger.info(f"📖 Fetching bug details for ID: {bug_id}")
        
        bug_details = fetch_bug_details(
            bug_id=int(bug_id),
            base_url=tfs_config.get('base_url', ''),
            username=tfs_config.get('username', ''),
            password=tfs_config.get('password', ''),
            pat=tfs_config.get('pat_token', '')
        )
        
        logger.info(f"✅ Bug details fetched successfully: ID={bug_id}, Title={bug_details.get('title', '')}")
        
        return {
            "success": True,
            "bug_details": bug_details,
            "timestamp": datetime.now().isoformat()
        }
    
    except Exception as e:
        logger.error(f"❌ Error fetching bug details: {str(e)}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }


@app.post("/api/tfs/update-bug")
async def update_bug_endpoint(request: dict):
    """
    Update an existing bug in TFS
    
    Request body:
    {
        "bug_id": 12345,
        "title": "Updated title",
        "reproduction_steps": "Updated steps",
        "severity": "2 - High",
        "priority": "1",
        "tfs_config": {...}
    }
    """
    try:
        from .tfs_tool import update_bug
        
        bug_id = request.get("bug_id")
        tfs_config = request.get("tfs_config")
        
        if not bug_id:
            return {
                "success": False,
                "error": "Bug ID is required",
                "timestamp": datetime.now().isoformat()
            }
        
        if not tfs_config:
            return {
                "success": False,
                "error": "TFS configuration is required",
                "timestamp": datetime.now().isoformat()
            }
        
        logger.info(f"🔄 Updating bug ID: {bug_id}")
        
        # Call update_bug with all provided fields
        response = update_bug(
            bug_id=int(bug_id),
            title=request.get("title"),
            description=request.get("description"),
            reproduction_steps=request.get("reproduction_steps"),
            severity=request.get("severity"),
            priority=request.get("priority"),
            assigned_to=request.get("assigned_to"),
            iteration_path=request.get("iteration_path"),
            area_path=request.get("area_path"),
            tags=request.get("tags"),
            base_url=tfs_config.get('base_url', ''),
            username=tfs_config.get('username', ''),
            password=tfs_config.get('password', ''),
            pat=tfs_config.get('pat_token', '')
        )
        
        if response.status_code in [200, 201]:
            logger.info(f"✅ Bug updated successfully: ID={bug_id}")
            return {
                "success": True,
                "bug_id": bug_id,
                "message": f"Bug {bug_id} updated successfully",
                "timestamp": datetime.now().isoformat()
            }
        else:
            logger.error(f"❌ Failed to update bug: {response.status_code} - {response.text}")
            return {
                "success": False,
                "error": f"TFS API error: {response.status_code}",
                "message": response.text[:500] if response.text else "Unknown error",
                "timestamp": datetime.now().isoformat()
            }
    
    except Exception as e:
        logger.error(f"❌ Error updating bug: {str(e)}", exc_info=True)
        return {
            "success": False,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }


@app.get("/api/testcase/default-sop")
async def get_default_sop():
    """Return default TruDocs SOP text for Agent 2."""
    try:
        from .prompts_manager import PromptsManager
        # Reload to ensure we get the latest from prompts.txt
        PromptsManager.reload_prompts()
        sop_text = PromptsManager.get_trudocs_sop()
        
        if not sop_text:
            return {
                "success": False,
                "sop_text": "",
                "message": "Default SOP not found in prompts.txt",
                "timestamp": datetime.now().isoformat(),
            }
            
        return {
            "success": True,
            "sop_text": sop_text,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "sop_text": "",
            "message": f"Failed to load default SOP: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/testcase/story-details")
async def fetch_testcase_story_details(request: UserStoryFetchRequest):
    """Fetch user story details for Agent 2 Step 2 UI preview."""
    try:
        from .tfs_tool import fetch_user_story_details

        cfg = request.tfs_config
        details = fetch_user_story_details(
            work_item_id=request.work_item_id,
            base_url=cfg.base_url if cfg else None,
            username=cfg.username if cfg else None,
            password=cfg.password if cfg else None,
            pat=cfg.pat_token if cfg else None,
        )
        return {
            "success": True,
            "story": details,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        return {
            "success": False,
            "story": {},
            "message": f"Failed to fetch story details: {str(e)}",
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/files/validate-drive-link")
async def validate_drive_link(request: DriveLinkValidationRequest):
    """Validate OneDrive/Google Drive link accessibility and return sheet names."""
    try:
        provider = (request.provider or "").strip().lower()
        file_url = (request.file_url or "").strip()

        if provider not in {"onedrive", "gdrive"}:
            return {
                "success": False,
                "accessible": False,
                "message": "Unsupported provider. Use 'onedrive' or 'gdrive'.",
                "sheet_names": [],
                "timestamp": datetime.now().isoformat(),
            }

        if not file_url:
            return {
                "success": False,
                "accessible": False,
                "message": "File URL is required.",
                "sheet_names": [],
                "timestamp": datetime.now().isoformat(),
            }

        token_from_env = ""
        if provider == "onedrive":
            token_from_env = os.getenv("ONEDRIVE_ACCESS_TOKEN", "").strip()
        elif provider == "gdrive":
            token_from_env = os.getenv("GOOGLE_DRIVE_ACCESS_TOKEN", "").strip()
        access_token = (request.access_token or token_from_env or "").strip()

        candidates = _build_download_candidates(provider, file_url)
        file_bytes, resolved_url = _download_excel_bytes(
            candidates,
            provider=provider,
            access_token=access_token,
            source_url=file_url,
        )
        sheet_names = _extract_sheet_names(file_bytes)

        if not sheet_names:
            return {
                "success": False,
                "accessible": False,
                "message": "File is accessible but no sheets were found.",
                "sheet_names": [],
                "timestamp": datetime.now().isoformat(),
            }

        return {
            "success": True,
            "accessible": True,
            "message": "Link is accessible and sheets loaded successfully.",
            "sheet_names": sheet_names,
            "resolved_url": resolved_url,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as ex:
        msg = str(ex)
        if "403" in msg:
            msg = "HTTP 403: Link access is blocked for direct download. Use a public share link or provide a valid access token."
        if "html page instead of a file" in msg.lower():
            msg = "The link opens a web/login page instead of file bytes. Use a direct downloadable share link or provide a valid access token."
        return {
            "success": False,
            "accessible": False,
            "message": f"Unable to access the file: {msg}",
            "sheet_names": [],
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/files/validate-excel-upload")
async def validate_excel_upload(file: UploadFile = File(...)):
    """Validate uploaded Excel/CSV and return sheet names."""
    try:
        content = await file.read()
        if not content:
            return {
                "success": False,
                "accessible": False,
                "message": "Uploaded file is empty.",
                "sheet_names": [],
                "timestamp": datetime.now().isoformat(),
            }

        if len(content) > 25 * 1024 * 1024:
            return {
                "success": False,
                "accessible": False,
                "message": "File is too large (max 25 MB for validation).",
                "sheet_names": [],
                "timestamp": datetime.now().isoformat(),
            }

        sheet_names = _extract_sheet_names(content)
        return {
            "success": True,
            "accessible": True,
            "message": "File loaded successfully.",
            "sheet_names": sheet_names,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as ex:
        return {
            "success": False,
            "accessible": False,
            "message": f"Unable to read uploaded file: {str(ex)}",
            "sheet_names": [],
            "timestamp": datetime.now().isoformat(),
        }


@app.post("/api/agent/execute/tfs-task/bulk-upload")
async def execute_tfs_task_bulk_upload(
    file: UploadFile = File(...),
    iteration_path: str = Form(...),
    sheet_name: Optional[str] = Form(None),
    tfs_base_url: Optional[str] = Form(None),
    tfs_username: Optional[str] = Form(None),
    tfs_password: Optional[str] = Form(None),
    tfs_pat_token: Optional[str] = Form(None),
    mode: Optional[str] = Form("create"), # Added mode parameter
):
    """Execute TFS task creation in batch mode from uploaded Excel file."""
    temp_path = None
    try:
        if not iteration_path or not iteration_path.strip():
            return {"status": "error", "error": "Iteration path is required"}

        content = await file.read()
        if not content:
            return {"status": "error", "error": "Uploaded file is empty"}

        suffix = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            temp_path = tmp.name

        from .agents.tfs_task_agent import execute_task_creation

        tfs_config = {
            "base_url": (tfs_base_url or "").strip(),
            "username": (tfs_username or "").strip(),
            "password": (tfs_password or "").strip(),
            "pat_token": (tfs_pat_token or "").strip(),
        }

        result = execute_task_creation(
            excel_file=temp_path,
            iteration_path=iteration_path.strip(),
            tfs_config=tfs_config,
            batch_mode=True,
            sheet_name=(sheet_name or "").strip() or None,
            mode=mode, # Pass mode
        )
        return JSONResponse(content=_safe_json_content(result))
    except Exception as ex:
        return JSONResponse(content=_safe_json_content({"status": "error", "error": str(ex)}))
    finally:
        if temp_path:
            try:
                os.remove(temp_path)
            except Exception:
                pass


@app.post("/api/agent/execute/tfs-task/bulk-drive")
async def execute_tfs_task_bulk_drive(request: DriveBulkTaskRequest):
    """Execute TFS task creation in batch mode from OneDrive/Google Drive link."""
    temp_path = None
    try:
        provider = (request.provider or "").strip().lower()
        if provider not in {"onedrive", "gdrive"}:
            return {"status": "error", "error": "Unsupported provider. Use 'onedrive' or 'gdrive'."}

        file_url = (request.file_url or "").strip()
        if not file_url:
            return {"status": "error", "error": "File URL is required"}

        iteration_path = (request.iteration_path or "").strip()
        if not iteration_path:
            return {"status": "error", "error": "Iteration path is required"}

        token_from_env = ""
        if provider == "onedrive":
            token_from_env = os.getenv("ONEDRIVE_ACCESS_TOKEN", "").strip()
        else:
            token_from_env = os.getenv("GOOGLE_DRIVE_ACCESS_TOKEN", "").strip()
        access_token = (request.access_token or token_from_env or "").strip()

        candidates = _build_download_candidates(provider, file_url)
        file_bytes, _resolved_url = _download_excel_bytes(
            candidates,
            provider=provider,
            access_token=access_token,
            source_url=file_url,
        )

        _extract_sheet_names(file_bytes)

        suffix = ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(file_bytes)
            temp_path = tmp.name

        from .agents.tfs_task_agent import execute_task_creation

        req_tfs = request.tfs_config
        tfs_config = {
            "base_url": ((req_tfs.base_url if req_tfs else "") or "").strip(),
            "username": ((req_tfs.username if req_tfs else "") or "").strip(),
            "password": ((req_tfs.password if req_tfs else "") or "").strip(),
            "pat_token": ((req_tfs.pat_token if req_tfs else "") or "").strip(),
        }

        result = execute_task_creation(
            excel_file=temp_path,
            iteration_path=iteration_path,
            tfs_config=tfs_config,
            batch_mode=True,
            sheet_name=(request.sheet_name or "").strip() or None,
            mode=request.mode or "create", # Pass mode
        )
        return JSONResponse(content=_safe_json_content(result))
    except Exception as ex:
        return JSONResponse(content=_safe_json_content({"status": "error", "error": str(ex)}))
    finally:
        if temp_path:
            try:
                os.remove(temp_path)
            except Exception:
                pass

# ==================== Agent Endpoints ====================
# Contains execution endpoints for all three agents plus background task handlers
# Each agent section includes: endpoint + background task runner

# ==================== AGENT 1: TFS Task Agent ====================
# Creates bulk TFS work items from various sources (Excel, OneDrive, Google Drive)
# Endpoints:
#   - POST /api/agent/execute/tfs-task (main execution)
#   - POST /api/agent/execute/tfs-task/bulk-upload (Excel upload)
#   - POST /api/agent/execute/tfs-task/bulk-drive (Drive integration)

@app.get("/api/agents")
async def get_available_agents():
    """Get list of available agents"""
    return {
        "agents": [
            {
                "id": "tfs_task",
                "name": "Agent #1: TFS Task Creation",
                "description": "Create and manage TFS task work items",
                "input_fields": [
                    {"name": "work_item_id", "type": "number", "label": "User Story ID", "required": True},
                    {"name": "task_description", "type": "text", "label": "Additional Context (Optional)", "required": False}
                ]
            },
            {
                "id": "testcase",
                "name": "Agent #2: Test Case Generation",
                "description": "Generate comprehensive test cases",
                "input_fields": [
                    {"name": "work_item_id", "type": "number", "label": "User Story ID", "required": True},
                    {"name": "sop_text", "type": "textarea", "label": "SOP Content", "required": False},
                    {"name": "test_mode", "type": "select", "label": "Test Type", "required": False},
                    {"name": "functional_prompt", "type": "textarea", "label": "Functional Prompt", "required": False},
                    {"name": "ui_prompt", "type": "textarea", "label": "UI Prompt", "required": False},
                ]
            }
        ]
    }

@app.post("/api/agent/execute/tfs-task")
# [AGENT 1] TFS Task Agent - Main Execution Endpoint
async def execute_tfs_task_agent(request: TaskCreationRequest, background_tasks: BackgroundTasks):
    """[AGENT 1] Execute TFS Task Agent - Create bulk work items from file/drive"""
    try:
        # Validate TFS config exists
        if not request.tfs_config:
            raise HTTPException(status_code=400, detail="TFS configuration is required")
        
        # Validate base_url is present
        base_url = (request.tfs_config.base_url or '').strip()
        if not base_url:
            raise HTTPException(status_code=400, detail="TFS Base URL is required")
        
        # Check for proper authentication: either PAT token or (username + password)
        pat_token = (request.tfs_config.pat_token or '').strip()
        username = (request.tfs_config.username or '').strip()
        password = (request.tfs_config.password or '').strip()
        
        if not pat_token and not (username and password):
            raise HTTPException(status_code=400, detail="Either PAT token OR (username + password) is required for TFS authentication")
        
        # Just log config info
        logger.info(f"🔧 TFS Task execution requested - Base URL: {base_url}")
        
        exec_id = f"exec_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
        
        active_executions[exec_id] = {
            "status": "running",
            "agent": "TFS Task Agent",
            "progress": 0,
            "start_time": datetime.now().isoformat()
        }
        
        config_dict = request.llm_config.dict() if request.llm_config else None
        tfs_config_dict = request.tfs_config.dict() if request.tfs_config else None
        
        background_tasks.add_task(
            run_task_agent,
            exec_id,
            request.work_item_id,
            request.task_description,
            config_dict,
            tfs_config_dict,
            request.batch_mode,
            request.excel_file,
            request.iteration_path,
            request.sheet_name
        )
        
        return {
            "execution_id": exec_id,
            "status": "started",
            "agent": "TFS Task Agent",
            "timestamp": datetime.now().isoformat()
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== AGENT 2: Test Case Agent ====================
# Generates comprehensive test cases (functional & UI) from work item details
# Endpoints:
#   - POST /api/agent/execute/testcase (main execution)
#   - POST /api/agent/review-test-cases (review generated tests)
#   - POST /api/agent/generate-missing-testcases (find gaps)
#   - POST /api/agent/testcase/analyze-story (story analysis)

@app.post("/api/agent/execute/testcase")
# [AGENT 2] Test Case Agent - Main Execution Endpoint
async def execute_testcase_agent(request: TestCaseGenerationRequest, background_tasks: BackgroundTasks):
    """[AGENT 2] Execute Test Case Agent - Generate functional & UI test cases"""
    try:
        if request.tfs_config or request.llm_config:
            logger.info(f"🗙 Test Case generation requested")
        
        exec_id = f"exec_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
        
        active_executions[exec_id] = {
            "status": "running",
            "agent": "Test Case Agent",
            "progress": 0,
            "start_time": datetime.now().isoformat()
        }
        
        # Use LLM config from request (from frontend user configuration)
        config_dict = request.llm_config.dict() if request.llm_config else None
        tfs_config_dict = request.tfs_config.dict() if request.tfs_config else None
        
        background_tasks.add_task(
            run_testcase_agent,
            exec_id,
            request.work_item_id,
            request.story_details,
            request.sop_text,
            config_dict,
            tfs_config_dict,
            request.test_mode,
            request.functional_prompt,
            request.ui_prompt,
            request.ui_screenshot_name,
            request.ui_screenshot_data,
            request.ui_screenshot_names or [],
            request.ui_screenshot_data_list or [],
            request.coverage_analysis if request.coverage_analysis is not None else False
        )
        
        return {
            "execution_id": exec_id,
            "status": "started",
            "agent": "Test Case Agent",
            "timestamp": datetime.now().isoformat()
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==================== AGENT 3: Bug Creation Agent ====================
# Creates TFS bug work items with comprehensive details and severity levels
# Endpoints:
#   - POST /api/agent/execute/bug-creation (main execution)
#   - POST /api/agent/create-bug (legacy endpoint)
#   - POST /api/agent/format-bug-report (report formatting)

@app.post("/api/agent/execute/bug-creation")
# [AGENT 3] Bug Creation Agent - Main Execution Endpoint
async def execute_bug_creation_agent(request: BugCreationRequest, background_tasks: BackgroundTasks):
    """
    Execute Bug Creation Agent with automatic self-healing code review
    (Self-heal is built-in and hidden - no UI control needed)
    """
    try:
        if request.tfs_config or request.llm_config:
            logger.info(f"🐛 Bug Creation execution requested")
        
        exec_id = f"exec_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
        
        active_executions[exec_id] = {
            "status": "running",
            "agent": "Bug Creation Agent (Auto Self-Heal)",
            "progress": 0,
            "start_time": datetime.now().isoformat()
        }
        
        config_dict = request.llm_config.dict() if request.llm_config else None
        tfs_config_dict = request.tfs_config.dict() if request.tfs_config else None
        
        background_tasks.add_task(
            run_bug_creation_agent,
            exec_id,
            request.work_item_id,
            request.bug_title,
            request.bug_description,
            request.reproduction_steps,
            request.expected_behavior,
            request.actual_behavior,
            request.severity,
            request.priority,
            config_dict,
            tfs_config_dict,
            request.found_in_version,
            request.assigned_to,
            request.work_item_type,
            request.is_update
        )
        
        return {
            "execution_id": exec_id,
            "status": "started",
            "agent": "Bug Creation Agent (with Auto Self-Heal)",
            "timestamp": datetime.now().isoformat()
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/execution/{exec_id}/status")
async def get_execution_status(exec_id: str):
    """Get execution status"""
    if exec_id not in active_executions:
        return {"status": "not_found", "message": f"Execution {exec_id} not found"}
    
    return active_executions[exec_id]

# ==================== Test Case Analysis & Review Endpoints ====================

@app.post("/api/agent/analyze-test-cases")
async def analyze_test_cases(request: TestCaseAnalysisRequest):
    """
    Analyze test cases based on user questions
    """
    try:
        logger.info(f"📊 Analyzing test cases with question: {request.question[:50]}...")
        
        from backend.agents.testcase_review_agent import execute_testcase_analysis
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        
        result = execute_testcase_analysis(
            test_cases=request.test_cases,
            story_details=request.story_details,
            question=request.question,
            chat_history=request.chat_history or [],
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Analysis completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Analysis error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Analysis failed: {str(e)}",
            "agent": "Test Case Analysis Agent"
        }


@app.post("/api/agent/review-test-cases")
async def review_test_cases(request: TestCaseReviewRequest):
    """
    Perform comprehensive review of test cases
    """
    try:
        logger.info("🔍 Starting test case review...")
        
        from backend.agents.testcase_review_agent import execute_testcase_review
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        
        result = execute_testcase_review(
            test_cases=request.test_cases,
            story_details=request.story_details,
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Review completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Review error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Review failed: {str(e)}",
            "agent": "Test Case Review Agent"
        }


@app.post("/api/agent/generate-missing-testcases")
async def generate_missing_testcases(request: dict):
    """
    Generate missing test cases based on review findings
    
    Request body:
    {
        "story_details": "user story",
        "review_text": "review findings with missing cases",
        "llm_config": {...}
    }
    """
    try:
        logger.info("📝 Generating missing test cases from review...")
        
        from backend.agents.testcase_review_agent import execute_generate_missing_testcases
        
        llm_config_dict = None
        if request.get('llm_config'):
            if isinstance(request.get('llm_config'), dict):
                llm_config_dict = request.get('llm_config')
        
        result = execute_generate_missing_testcases(
            story_details=request.get('story_details', ''),
            review_text=request.get('review_text', ''),
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Missing test cases generation completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Missing test cases generation error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Generation failed: {str(e)}",
            "agent": "Missing Test Case Generator"
        }

@app.post("/api/agent/testcase/analyze-story")
async def analyze_story(request: StoryAnalysisRequest):
    """
    Analyze a User Story from a tester's perspective.
    """
    try:
        logger.info("🧪 Analyzing User Story for strategic testing breakdown...")
        
        from .prompts_manager import PromptsManager
        from .llm_config import get_configured_llm
        
        prompt_template = PromptsManager.get_story_analysis_prompt()
        if not prompt_template:
            # Fallback if specific tag not found
            prompt_template = "Act as a Senior QA Analyst. Analyze this User Story and provide a testing breakdown: {story_details}"
        
        # We need to make sure {story_details} is in the template or handle it.
        if "{story_details}" in prompt_template:
            prompt = prompt_template.replace("{story_details}", request.story_text)
        else:
            prompt = f"{prompt_template}\n\n### USER STORY DETAILS:\n{request.story_text}"
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        llm = get_configured_llm(llm_config_dict)
        
        response = llm.call([{"role": "user", "content": prompt}])
        
        return {
            "status": "success",
            "analysis": response,
            "agent": "Story Analyzer Agent"
        }
    
    except Exception as e:
        logger.error(f"❌ Story Analysis error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Analysis failed: {str(e)}",
            "agent": "Story Analyzer Agent"
        }


# ==================== Background Tasks ====================

@app.get("/api/execution/{exec_id}/status")
async def get_execution_status(exec_id: str):
    """Get execution status"""
    if exec_id not in active_executions:
        return {"status": "not_found", "message": f"Execution {exec_id} not found"}
    
    return active_executions[exec_id]

# ==================== Test Case Analysis & Review Endpoints ====================

@app.post("/api/agent/analyze-test-cases")
async def analyze_test_cases(request: TestCaseAnalysisRequest):
    """
    Analyze test cases based on user questions
    """
    try:
        logger.info(f"📊 Analyzing test cases with question: {request.question[:50]}...")
        
        from backend.agents.testcase_review_agent import execute_testcase_analysis
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        
        result = execute_testcase_analysis(
            test_cases=request.test_cases,
            story_details=request.story_details,
            question=request.question,
            chat_history=request.chat_history or [],
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Analysis completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Analysis error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Analysis failed: {str(e)}",
            "agent": "Test Case Analysis Agent"
        }


@app.post("/api/agent/review-test-cases")
async def review_test_cases(request: TestCaseReviewRequest):
    """
    Perform comprehensive review of test cases
    """
    try:
        logger.info("🔍 Starting test case review...")
        
        from backend.agents.testcase_review_agent import execute_testcase_review
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        
        result = execute_testcase_review(
            test_cases=request.test_cases,
            story_details=request.story_details,
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Review completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Review error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Review failed: {str(e)}",
            "agent": "Test Case Review Agent"
        }


@app.post("/api/agent/generate-missing-testcases")
async def generate_missing_testcases(request: dict):
    """
    Generate missing test cases based on review findings
    
    Request body:
    {
        "story_details": "user story",
        "review_text": "review findings with missing cases",
        "llm_config": {...}
    }
    """
    try:
        logger.info("📝 Generating missing test cases from review...")
        
        from backend.agents.testcase_review_agent import execute_generate_missing_testcases
        
        llm_config_dict = None
        if request.get('llm_config'):
            if isinstance(request.get('llm_config'), dict):
                llm_config_dict = request.get('llm_config')
        
        result = execute_generate_missing_testcases(
            story_details=request.get('story_details', ''),
            review_text=request.get('review_text', ''),
            llm_config=llm_config_dict
        )
        
        logger.info(f"✅ Missing test cases generation completed: {result.get('status')}")
        return result
    
    except Exception as e:
        logger.error(f"❌ Missing test cases generation error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Generation failed: {str(e)}",
            "agent": "Missing Test Case Generator"
        }

@app.post("/api/agent/testcase/chat-story")
async def chat_story(request: StoryChatRequest):
    """
    Continue conversation about a User Story analysis.
    """
    try:
        logger.info(f"💬 Continuing conversation about User Story analysis...")
        
        from .llm_config import get_configured_llm
        
        messages = []
        # Initial context
        messages.append({"role": "system", "content": "You are a Senior QA Analyst. You have already provided an initial analysis of a User Story. Now, answer follow-up questions from the user based on that story details."})
        messages.append({"role": "system", "content": f"USER STORY DETAILS:\n{request.story_text}"})
        
        # Add chat history
        if request.chat_history:
            for msg in request.chat_history:
                messages.append(msg)
        
        # Add current question
        messages.append({"role": "user", "content": request.question})
        
        llm_config_dict = request.llm_config.dict() if request.llm_config else None
        llm = get_configured_llm(llm_config_dict)
        
        response = llm.call(messages)
        
        return {
            "status": "success",
            "reply": response,
            "agent": "Story Chat Agent"
        }
    
    except Exception as e:
        logger.error(f"❌ Story Chat error: {str(e)}", exc_info=True)
        return {
            "status": "error",
            "error": f"Chat failed: {str(e)}",
            "agent": "Story Chat Agent"
        }


def apply_self_healing(result: dict, llm_config: Optional[Dict] = None, agent_name: str = "Agent") -> dict:
    """
    Apply self-healing (code review) to agent output without modifying agent code
    
    Args:
        result: Agent execution result dict with 'status' and 'result' keys
        llm_config: Optional LLM configuration for self-healing
        agent_name: Name of the agent for logging
    
    Returns:
        Modified result dict with reviewed content
    """
    if not llm_config or not result.get("result"):
        return result
    
    try:
        from backend.llm_config import get_llm_client
        client = get_llm_client(llm_config)
        
        # Self-heal prompt focuses on quality, not modification
        heal_prompt = f"""Review this {agent_name} output and provide constructive feedback:
{result.get('result', '')}

Provide a brief professional review (max 3 sentences)."""
        
        response = client.create_message(heal_prompt)
        return {
            **result,
            "reviewed": response
        }
    except Exception as e:
        logger.debug(f"ℹ️ Self-healing not available: {str(e)}")
        return result
        if temp_dir and os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except Exception:
                pass



# ==================== Background Tasks ====================

def apply_self_healing(result: dict, llm_config: Optional[Dict] = None, agent_name: str = "Agent") -> dict:
    """
    Apply self-healing (code review) to agent output without modifying agent code
    
    Args:
        result: Agent execution result dict with 'status' and 'result' keys
        llm_config: Optional LLM configuration for self-healing
        agent_name: Name of the agent for logging
        
    Returns:
        Updated result dict with self-healing applied
    """
    if not llm_config:
        logger.debug(f"ℹ️ Self-healing skipped for {agent_name} (no LLM config)")
        return result
    
    if result.get("status") != "success":
        logger.debug(f"ℹ️ Self-healing skipped for {agent_name} (agent returned {result.get('status')} status)")
        return result
    
    try:
        from .agents.code_reviewer_agent import execute_code_review
        
        content_to_review = result.get("result", "")
        if not content_to_review:
            logger.debug(f"ℹ️ Self-healing skipped for {agent_name} (no content to review)")
            return result
        
        logger.debug(f"🔧 Applying self-healing to {agent_name} output...")
        review_result = execute_code_review(str(content_to_review), llm_config)
        
        if review_result and review_result.get("status") != "error":
            result["reviewed"] = review_result
            if review_result.get("status") == "fixed":
                result["result"] = review_result.get("result", content_to_review)
                logger.info(f"✨ {agent_name} output auto-fixed by self-healing")
            else:
                logger.debug(f"✓ {agent_name} output validated as correct")
        else:
            logger.debug(f"⚠️ Self-healing review indicated issues for {agent_name}")
            
    except Exception as e:
        logger.debug(f"ℹ️ Self-healing failed for {agent_name} (optional): {str(e)}")
    
    return result

# ==================== Background Tasks ====================
# Background task runners for each agent (executed asynchronously)

# [AGENT 1] TFS Task Agent - Background Task
def run_task_agent(exec_id: str, work_item_id: int, task_description: str, llm_config: Optional[Dict], tfs_config: Optional[Dict] = None, batch_mode: bool = False, excel_file: Optional[str] = None, iteration_path: Optional[str] = None, sheet_name: Optional[str] = None):
    """Background task: Run TFS Task Agent"""
    try:
        from .agents.tfs_task_agent import execute_task_creation
        from .agents.code_reviewer_agent import execute_code_review
        
        active_executions[exec_id]["progress"] = 25
        
        result = execute_task_creation(
            work_item_id=work_item_id,
            task_description=task_description,
            llm_config=llm_config,
            tfs_config=tfs_config,
            batch_mode=batch_mode,
            excel_file=excel_file,
            iteration_path=iteration_path,
            sheet_name=sheet_name
        )
        
        active_executions[exec_id]["progress"] = 75
        
        # Apply self-healing (built-in, non-intrusive)
        result = apply_self_healing(result, llm_config, "TFS Task Agent")
        active_executions[exec_id]["progress"] = 90
        
        active_executions[exec_id]["status"] = "completed"
        active_executions[exec_id]["progress"] = 100
        active_executions[exec_id]["result"] = result
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        
        if "default" not in execution_history:
            execution_history["default"] = []
        execution_history["default"].append({
            "execution_id": exec_id,
            "agent": "TFS Task Agent",
            "status": result["status"],
            "timestamp": datetime.now().isoformat(),
            "work_item_id": work_item_id
        })
    
    except Exception as e:
        error_msg = str(e)
        print(f"❌ run_task_agent exception: {error_msg}")
        import traceback
        traceback.print_exc()
        
        active_executions[exec_id]["status"] = "error"
        active_executions[exec_id]["error"] = error_msg
        active_executions[exec_id]["progress"] = 100
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        active_executions[exec_id]["result"] = {
            "status": "error",
            "error": error_msg,
            "summary": {"created": 0, "failed": 0, "updated": 0, "total": 0},
            "created_ids": [],
            "updated_ids": [],
            "report_rows": [],
            "errors": [error_msg],
            "agent": "TFS Task Agent (Bulk)"
        }

# [AGENT 2] Test Case Agent - Background Task
def run_testcase_agent(
    exec_id: str,
    work_item_id: Optional[int],
    story_details: Optional[str],
    sop_text: str,
    llm_config: Optional[Dict],
    tfs_config: Optional[Dict] = None,
    test_mode: str = "functional",
    functional_prompt: str = "",
    ui_prompt: str = "",
    ui_screenshot_name: str = "",
    ui_screenshot_data: str = "",
    ui_screenshot_names: Optional[List[str]] = None,
    ui_screenshot_data_list: Optional[List[str]] = None,
    coverage_analysis: bool = False,
):
    """Background task: Run Test Case Agent"""
    try:
        from .agents.testcase_agent import execute_testcase_generation
        from .agents.code_reviewer_agent import execute_code_review
        
        active_executions[exec_id]["progress"] = 25
        
        result = execute_testcase_generation(
            work_item_id=work_item_id,
            story_details=story_details,
            sop_text=sop_text,
            llm_config=llm_config,
            tfs_config=tfs_config,
            test_mode=test_mode,
            functional_prompt=functional_prompt,
            ui_prompt=ui_prompt,
            ui_screenshot_name=ui_screenshot_name,
            ui_screenshot_data=ui_screenshot_data,
            ui_screenshot_names=ui_screenshot_names or [],
            ui_screenshot_data_list=ui_screenshot_data_list or [],
            coverage_analysis=coverage_analysis,
        )
        
        active_executions[exec_id]["progress"] = 75
        
        # Apply self-healing (built-in, non-intrusive)
        result = apply_self_healing(result, llm_config, "Test Case Agent")
        active_executions[exec_id]["progress"] = 90
        
        active_executions[exec_id]["status"] = "completed"
        active_executions[exec_id]["progress"] = 100
        active_executions[exec_id]["result"] = result
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        
        if "default" not in execution_history:
            execution_history["default"] = []
        execution_history["default"].append({
            "execution_id": exec_id,
            "agent": "Test Case Agent",
            "status": result["status"],
            "timestamp": datetime.now().isoformat(),
            "work_item_id": work_item_id
        })

    except Exception as e:
        active_executions[exec_id]["status"] = "error"
        active_executions[exec_id]["error"] = str(e)
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()

# [AGENT 3] Bug Creation Agent - Background Task
def run_bug_creation_agent(
    exec_id: str,
    work_item_id: Optional[int],
    bug_title: str,
    bug_description: str,
    reproduction_steps: str,
    expected_behavior: str,
    actual_behavior: str,
    severity: str,
    priority: str,
    llm_config: Optional[Dict],
    tfs_config: Optional[Dict] = None,
    found_in_version: str = "",
    assigned_to: str = "",
    work_item_type: str = "Bug",
    is_update: bool = False,
):
    """Background task: Run Bug & Feature Creation Agent"""
    temp_files = []
    try:
        from .agents.bug_creation_agent import execute_bug_creation
        from .agents.code_reviewer_agent import execute_code_review
        
        wi_type = work_item_type or "Bug"
        active_executions[exec_id]["progress"] = 25
        active_executions[exec_id]["agent"] = f"{wi_type} Creation Agent"
        
        # Step 1: Execute creation
        result = execute_bug_creation(
            work_item_id=work_item_id,
            work_item_type=wi_type,
            bug_description=bug_description,
            bug_title=bug_title,
            reproduction_steps=reproduction_steps,
            expected_behavior=expected_behavior,
            actual_behavior=actual_behavior,
            severity=severity,
            priority=priority,
            llm_config=llm_config,
            tfs_config=tfs_config,
            found_in_version=found_in_version,
            assigned_to=assigned_to,
            is_update=is_update,
        )
        
        active_executions[exec_id]["progress"] = 75
        
        # Apply self-healing (built-in, non-intrusive)
        # For bug creation, we review the bug details, not the full result
        if llm_config and (result.get("success") or result.get("created_count", 0) > 0):
            try:
                bug_details = f"""
Bug Title: {bug_title}
Reproduction Steps: {reproduction_steps}
Expected Behavior: {expected_behavior}
Actual Behavior: {actual_behavior}
Severity: {severity}
Priority: {priority}
"""
                result = apply_self_healing(
                    {"status": "success", "result": bug_details}, 
                    llm_config, 
                    "Bug Creation Agent"
                )
            except Exception as e:
                logger.debug(f"ℹ️ Self-healing failed for bug details: {str(e)}")
        
        active_executions[exec_id]["progress"] = 90
        
        active_executions[exec_id]["status"] = "completed"
        active_executions[exec_id]["progress"] = 100
        active_executions[exec_id]["result"] = result
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        
        if "default" not in execution_history:
            execution_history["default"] = []
        execution_history["default"].append({
            "execution_id": exec_id,
            "agent": "Bug Creation Agent (with Self-Heal)",
            "status": result.get("status", "success"),
            "timestamp": datetime.now().isoformat(),
            "work_item_id": work_item_id,
            "bugs_created": result.get("created_count", 0)
        })
    
    except Exception as e:
        active_executions[exec_id]["status"] = "error"
        active_executions[exec_id]["error"] = str(e)
        active_executions[exec_id]["end_time"] = datetime.now().isoformat()
        logger.error(f"❌ Bug creation agent error: {str(e)}", exc_info=True)
        for f in temp_files:
            try:
                if os.path.exists(f): os.remove(f)
            except Exception: pass


# ==================== Dashboard Agent Endpoints (Agent #4) ====================

@app.post("/api/dashboard/queries")
async def dashboard_fetch_queries(request: DashboardQueriesRequest):
    """
    [AGENT 4] Fetch the list of saved TFS queries for the configured project.
    Returns: [{"id": "...", "name": "...", "path": "..."}]
    """
    try:
        tfs = request.tfs_config
        from .tfs_tool import _normalize_tfs_url_for_api

        # Prefer task_url (has project in path); fall back to base_url
        raw_url = ""
        if tfs:
            raw_url = (tfs.task_url or "").strip() or (tfs.base_url or "").strip()

        project_url = _normalize_tfs_url_for_api(raw_url).rstrip("/") if raw_url else ""

        if not project_url:
            raise HTTPException(status_code=400,
                detail="TFS URL is required — set Base URL or Task URL in TFS config")

        pat = (tfs.pat_token or "").strip() if tfs else ""
        username = (tfs.username or "").strip() if tfs else ""
        password = (tfs.password or "") if tfs else ""
        if not pat and not (username and password):
            raise HTTPException(status_code=400, detail="Dashboard Agent requires authentication: PAT or Username/Password")

        logger.info(f"Dashboard: fetching saved queries for {project_url}")

        from .agents.dashboard_agent import fetch_tfs_saved_queries
        import asyncio
        queries = await asyncio.get_event_loop().run_in_executor(
            None, fetch_tfs_saved_queries, project_url, pat, username, password
        )
        logger.info(f"Dashboard: found {len(queries)} queries")
        return {"status": "success", "queries": queries}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard queries fetch error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/dashboard/generate")
async def dashboard_generate(request: DashboardGenerateRequest):
    """
    [AGENT 4] Generate QA Activity Dashboard.
    mode=static  → tables + trend data (no LLM)
    mode=ai      → static data + LLM strategic narrative
    """
    try:
        tfs = request.tfs_config
        from .tfs_tool import _normalize_tfs_url_for_api

        raw_url = ""
        if tfs:
            raw_url = (tfs.task_url or "").strip() or (tfs.base_url or "").strip()

        project_url = _normalize_tfs_url_for_api(raw_url).rstrip("/") if raw_url else ""

        if not project_url:
            raise HTTPException(status_code=400,
                detail="TFS URL is required — set Base URL or Task URL in TFS config")

        pat = (tfs.pat_token or "").strip() if tfs else ""

        def _b64_to_bytes(b64: Optional[str]) -> Optional[bytes]:
            if not b64:
                return None
            try:
                return base64.b64decode(b64)
            except Exception:
                return None

        llm_config_dict = request.llm_config.dict() if request.llm_config else None

        from .agents.dashboard_agent import execute_dashboard_agent
        import asyncio
        
        logger.info(f"📥 Dashboard generate request received. "
                    f"VT size: {len(request.vertical_excel_b64) if request.vertical_excel_b64 else 0}, "
                    f"Auto size: {len(request.automation_excel_b64) if request.automation_excel_b64 else 0}, "
                    f"Perf size: {len(request.performance_excel_b64) if request.performance_excel_b64 else 0}")

        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            execute_dashboard_agent,
            project_url,
            tfs.dict() if tfs else {},
            request.bug_query_id or "",
            request.retest_query_id or "",
            request.story_query_id or "",
            request.other_query_id or "",
            _b64_to_bytes(request.vertical_excel_b64),
            _b64_to_bytes(request.automation_excel_b64),
            _b64_to_bytes(request.performance_excel_b64),
            request.mode or "static",
            request.llm_prompt or "",
            llm_config_dict,
        )

        logger.info(f"✅ Dashboard generation completed (mode={request.mode})")
        return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard generate error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    # Start the server on 0.0.0.0 so it's accessible from other machines in the company
    uvicorn.run(app, host="0.0.0.0", port=8000)
